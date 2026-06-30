#!/usr/bin/env python3
from __future__ import annotations

import copy
import dataclasses
import decimal
import re
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import dgm_database
import dgm_xlsx_common
import dgm_xlsx_preprocessor
from dgm_gui_common import openpyxl

if openpyxl is not None:
	import openpyxl.styles
	import openpyxl.utils
	from openpyxl.cell.cell import MergedCell
else:
	MergedCell = object  # type: ignore[assignment,misc]

DOCUMENT_TYPE_PRESENT = "present"
DOCUMENT_TYPE_MISSING = "missing"
INFORMATION_MISSING_TEXT = "Інформація відсутня"
FORMULARY_LABEL = "Вміст ДГМ згідно Формуляру:"
MISSING_LABEL = "Вміст ДГМ, яких не вистачає:"
TOTAL_IN_PRODUCT_LABEL = "Всього у виробі:"
HEADER_END_MARKER = "Підрахунок здійснили"
CIVILIAN_RANK = "працівник ЗСУ"
OPTIONAL_DOT = r"\.?"
RANK_WORD_SEPARATOR = r"(?:\s+|\s*-\s*)?"
RANK_COMPACT_SEPARATOR = r"[-\s]*"
RANK_REQUIRED_SEPARATOR = r"(?:\s+|\s*-\s*)"

JUNIOR_PREFIX_RE = rf"(?:молодший|мол{OPTIONAL_DOT}|мл{OPTIONAL_DOT})"
SENIOR_PREFIX_RE = rf"(?:старший|ст{OPTIONAL_DOT})"
CHIEF_PREFIX_RE = rf"(?:головний|гол{OPTIONAL_DOT})"

SERGEANT_RE = rf"(?:сержант|серж{OPTIONAL_DOT}|с-нт{OPTIONAL_DOT}|с-т{OPTIONAL_DOT})"
MASTER_RE = r"(?:майстер|майстр|майст(?:е)?р)"
LIEUTENANT_RE = rf"(?:лейтенант|лейт{OPTIONAL_DOT}|л-нт|лнт|л-т)"
CAPTAIN_RE = rf"(?:капітан|кап{OPTIONAL_DOT}|к-н{OPTIONAL_DOT})"
GENERAL_RE = rf"(?:генерал|ген{OPTIONAL_DOT})"
ADMIRAL_RE = rf"(?:адмірал|адм{OPTIONAL_DOT})"

RANK_ALIAS_PATTERNS: Sequence[Tuple[str, str]] = (
	# Рядовий склад — армійські
	(r"рекр(?:ут)?", "рекрут"),
	(r"солд(?:ат)?\.?|ряд(?:овий)?\.?", "солдат"),
	(rf"{SENIOR_PREFIX_RE}\s*солд(?:ат)?\.?", "старший солдат"),

	# Рядовий склад — корабельні
	(r"матр(?:ос)?\.?", "матрос"),
	(rf"{SENIOR_PREFIX_RE}\s*матр(?:ос)?\.?", "старший матрос"),

	# Молодший сержантський / старшинський склад
	(rf"{JUNIOR_PREFIX_RE}\s*{SERGEANT_RE}|м/с-т", "молодший сержант"),
	(SERGEANT_RE, "сержант"),
	(r"(?:старшина|ст-на)\s+(?:2|ii)\s+ст(?:атті)?\.?", "старшина 2 статті"),
	(r"(?:старшина|ст-на)\s+(?:1|i)\s+ст(?:атті)?\.?", "старшина 1 статті"),

	# Старший сержантський / старшинський склад
	(rf"{SENIOR_PREFIX_RE}\s*{SERGEANT_RE}", "старший сержант"),
	(rf"{CHIEF_PREFIX_RE}\s*{SERGEANT_RE}", "головний сержант"),
	(rf"(?:штаб|шт{OPTIONAL_DOT}){RANK_WORD_SEPARATOR}{SERGEANT_RE}", "штаб-сержант"),
	(rf"{CHIEF_PREFIX_RE}\s*старшина", "головний старшина"),
	(rf"{CHIEF_PREFIX_RE}\s*корабельний\s+старшина", "головний корабельний старшина"),
	(rf"(?:штаб|шт{OPTIONAL_DOT}){RANK_WORD_SEPARATOR}старшина", "штаб-старшина"),

	# Вищий сержантський / старшинський склад
	(rf"{MASTER_RE}{RANK_REQUIRED_SEPARATOR}{SERGEANT_RE}", "майстер-сержант"),
	(rf"{SENIOR_PREFIX_RE}{RANK_REQUIRED_SEPARATOR}{MASTER_RE}{RANK_REQUIRED_SEPARATOR}{SERGEANT_RE}", "старший майстер-сержант"),
	(rf"{CHIEF_PREFIX_RE}{RANK_REQUIRED_SEPARATOR}{MASTER_RE}{RANK_REQUIRED_SEPARATOR}{SERGEANT_RE}", "головний майстер-сержант"),
	(rf"{MASTER_RE}{RANK_REQUIRED_SEPARATOR}старшина", "майстер-старшина"),
	(rf"{SENIOR_PREFIX_RE}{RANK_REQUIRED_SEPARATOR}{MASTER_RE}{RANK_REQUIRED_SEPARATOR}старшина", "старший майстер-старшина"),
	(rf"{CHIEF_PREFIX_RE}{RANK_REQUIRED_SEPARATOR}{MASTER_RE}{RANK_REQUIRED_SEPARATOR}старшина", "головний майстер-старшина"),

	# Молодший офіцерський склад
	(rf"{JUNIOR_PREFIX_RE}\s*{LIEUTENANT_RE}|м/л-нт", "молодший лейтенант"),
	(LIEUTENANT_RE, "лейтенант"),
	(rf"{SENIOR_PREFIX_RE}\s*{LIEUTENANT_RE}", "старший лейтенант"),
	(CAPTAIN_RE, "капітан"),
	(rf"{CAPTAIN_RE}{RANK_COMPACT_SEPARATOR}{LIEUTENANT_RE}", "капітан-лейтенант"),

	# Старший офіцерський склад
	(r"майор|м-р|мр", "майор"),
	(r"підполковник|підп-?к|п/п-?к|пп-?к", "підполковник"),
	(r"полковник|полк\.?|п-к|пк", "полковник"),
	(rf"{CAPTAIN_RE}\s+(?:3|iii)\s+(?:рангу|р\.?)", "капітан 3 рангу"),
	(rf"{CAPTAIN_RE}\s+(?:2|ii)\s+(?:рангу|р\.?)", "капітан 2 рангу"),
	(rf"{CAPTAIN_RE}\s+(?:1|i)\s+(?:рангу|р\.?)", "капітан 1 рангу"),

	# Вищий офіцерський склад
	(rf"(?:бригадний|бриг{OPTIONAL_DOT})\s*{GENERAL_RE}", "бригадний генерал"),
	(rf"{GENERAL_RE}{RANK_COMPACT_SEPARATOR}(?:майор|м-р)", "генерал-майор"),
	(rf"{GENERAL_RE}{RANK_COMPACT_SEPARATOR}{LIEUTENANT_RE}", "генерал-лейтенант"),
	(GENERAL_RE, "генерал"),
	(r"комм?одор", "коммодор"),
	(rf"контр[ -]{ADMIRAL_RE}|к\.?-?\s*адмірал", "контр-адмірал"),
	(rf"віце[ -]{ADMIRAL_RE}|в\.?-?\s*адмірал", "віце-адмірал"),
	(ADMIRAL_RE, "адмірал"),

	# Старі звання, які вже не присвоюються в цій системі, але можуть траплятись у документах
	(rf"{GENERAL_RE}{RANK_COMPACT_SEPARATOR}полковник", "генерал-полковник"),
	(rf"{GENERAL_RE}\s+армії(?:\s+україни)?", "генерал армії україни"),
)
RANK_ALIASES: Sequence[Tuple[re.Pattern[str], str]] = tuple(
	(re.compile(rf"^(?:{Pattern})$", re.IGNORECASE), CanonicalRank)
	for Pattern, CanonicalRank in RANK_ALIAS_PATTERNS
)

RANK_PREFIX_ALIASES: Sequence[Tuple[re.Pattern[str], str]] = tuple(
	(re.compile(rf"^(?:{Pattern})\s+(?P<name>.+)$", re.IGNORECASE), CanonicalRank)
	for Pattern, CanonicalRank in sorted(RANK_ALIAS_PATTERNS, key=lambda RankAlias: len(RankAlias[0]), reverse=True)
)
CANONICAL_FILENAME_RE = re.compile(r"^\s*(?:(?P<number>\d+)\.\s*)?Відомість\s+(?P<kind>не?комплектності|комплектності)\s+(?P<rest>.+?)\s*$", re.IGNORECASE)
YEAR_RE = re.compile(r"(?P<year>\b(?:19|20)\d{2}\b)\s*(?:року|рік|р\.)?\s*$", re.IGNORECASE)
SERIAL_RE = re.compile(r"(?:№|N\.?)\s*(?P<serial>.*?)\s*$", re.IGNORECASE)
FILE_NUMBER_RE = re.compile(r"^\s*(?P<number>\d+)\.\s+")
INVALID_FILENAME_CHAR_RE = re.compile(r"[<>:\"/\\|?*\x00-\x1f]")
VALID_SIMPLE_DECIMAL_NUMBER_RE = re.compile(r"^\d+(?:[,.]\d+)?$")
DGM_NUMBER_FORMAT = "0.############################;-0.############################;0"
DGM_BORDER_STYLE = "thin"
DGM_BORDER_COLOR = "000000"
MIN_COLUMN_WIDTH = 4.0
MAX_TEXT_COLUMN_WIDTH = 48.0
MAX_NUMERIC_COLUMN_WIDTH = 26.0
WIDTH_PADDING = 2.0


@dataclasses.dataclass
class FooterPerson:
	Rank: str
	Name: str


@dataclasses.dataclass
class DgmDocumentMetadata:
	FileNumber: Optional[int] = None
	DocumentType: Optional[str] = None
	EquipmentName: str = ""
	SerialNumber: str = "№б/н"
	ManufactureYear: str = ""
	People: str = ""
	FooterPeople: List[FooterPerson] = dataclasses.field(default_factory=list)
	FormularyValues: Optional[dgm_database.DgmValues] = None
	Sources: Dict[str, Dict[str, object]] = dataclasses.field(default_factory=dict)
	Conflicts: List[str] = dataclasses.field(default_factory=list)

	def CanonicalTitle(self) -> str:
		Kind = "некомплектності" if self.DocumentType == DOCUMENT_TYPE_MISSING else "комплектності"
		Parts = ["Відомість", Kind]
		if self.EquipmentName:
			Parts.append(self.EquipmentName)
		Parts.append(CanonicalSerial(self.SerialNumber))
		if self.ManufactureYear:
			Parts.extend([self.ManufactureYear, "року"])
		return " ".join(Parts)

	def CanonicalFilename(self) -> str:
		Prefix = f"{self.FileNumber}. " if self.FileNumber is not None else ""
		return f"{Prefix}{SanitizeFilenameText(self.CanonicalTitle())}.xlsx"


@dataclasses.dataclass
class PostprocessRowIssue:
	SheetName: str
	Row: int
	Name: str
	Reason: str


@dataclasses.dataclass
class PostprocessResult:
	OriginalPath: Path
	SavedPath: Path
	Metadata: DgmDocumentMetadata
	Issues: List[PostprocessRowIssue]
	Warnings: List[str]


class FooterPlacementRequired(RuntimeError):
	def __init__(self, FilePath: Path, ReviewStart: int) -> None:
		super().__init__(f"Cannot detect valid footer in '{FilePath}'. Manual footer row selection is required.")
		self.FilePath = FilePath
		self.ReviewStart = ReviewStart


def NormalizeSpaces(Text: object) -> str:
	return " ".join(str(Text or "").strip().split())


def ExpandRank(Text: object) -> str:
	Rank = NormalizeSpaces(Text).casefold().replace("–", "-").replace("—", "-")
	for RankPattern, CanonicalRank in RANK_ALIASES:
		if RankPattern.fullmatch(Rank):
			return CanonicalRank
	return CIVILIAN_RANK


def ParseFooterPeople(Text: object) -> List[FooterPerson]:
	Raw = NormalizeSpaces(Text)
	if HEADER_END_MARKER in Raw:
		Raw = Raw.split(HEADER_END_MARKER, 1)[1]
	Raw = Raw.strip(" :-–—")
	if not Raw:
		return []
	RankLookahead = "|".join(rf"(?:{Pattern})" for Pattern, _CanonicalRank in RANK_ALIAS_PATTERNS)
	Raw = re.sub(rf"\s+(?:та|і)\s+(?={RankLookahead}\b)", "; ", Raw, flags=re.IGNORECASE)
	Parts = [Part.strip(" .") for Part in re.split(r"[;,]", Raw) if Part.strip(" .")]
	People: List[FooterPerson] = []
	for Part in Parts:
		Rank = CIVILIAN_RANK
		Name = NormalizeSpaces(Part)
		for RankPattern, CanonicalRank in RANK_PREFIX_ALIASES:
			Match = RankPattern.match(Part)
			if Match:
				Rank = CanonicalRank
				Name = NormalizeSpaces(Match.group("name"))
				break
		if Name:
			People.append(FooterPerson(Rank, Name))
	return People

def SanitizeFilenameText(Text: object) -> str:
	Value = INVALID_FILENAME_CHAR_RE.sub("", NormalizeSpaces(Text))
	Value = re.sub(r"\s+", " ", Value).strip(" .")
	return Value or "document"


def FilenameCompatibleText(ExpectedText: object, FilenameText: object) -> bool:
	Expected = NormalizeSpaces(ExpectedText)
	Filename = NormalizeSpaces(FilenameText)
	PatternParts = []
	for Character in Expected:
		if INVALID_FILENAME_CHAR_RE.fullmatch(Character):
			PatternParts.append(".?")
		else:
			PatternParts.append(re.escape(Character))
	Pattern = "".join(PatternParts)
	return re.fullmatch(Pattern, Filename, flags=re.IGNORECASE) is not None

def CanonicalSerial(Text: object) -> str:
	Value = NormalizeSpaces(Text)
	Value = re.sub(r"^(?:№|N\.?)\s*", "", Value, flags=re.IGNORECASE).strip()
	if not Value:
		return "№б/н"
	return "№" + Value.replace(" ", "")


def ParseDocumentText(Text: str) -> Dict[str, object]:
	Name = Path(Text).stem if Text.casefold().endswith(".xlsx") else Text
	Match = CANONICAL_FILENAME_RE.match(Name)
	if not Match:
		return {}
	Rest = NormalizeSpaces(Match.group("rest"))
	Year = ""
	YearMatch = YEAR_RE.search(Rest)
	if YearMatch:
		Year = YearMatch.group("year")
		Rest = Rest[:YearMatch.start()].strip()
	Serial = "№б/н"
	SerialMatch = SERIAL_RE.search(Rest)
	if SerialMatch:
		Serial = CanonicalSerial(SerialMatch.group(0))
		Equipment = Rest[:SerialMatch.start()].strip()
	else:
		Equipment = Rest.strip()
	Kind = Match.group("kind").casefold()
	DocumentType = DOCUMENT_TYPE_MISSING if "некомплект" in Kind else DOCUMENT_TYPE_PRESENT
	Number = Match.group("number")
	return {
		"FileNumber": int(Number) if Number else None,
		"DocumentType": DocumentType,
		"EquipmentName": Equipment,
		"SerialNumber": Serial,
		"ManufactureYear": Year,
	}


def ParseRow1Number(Text: object) -> Optional[int]:
	Match = re.search(r"Відомість\s*№\s*(\d+)", NormalizeSpaces(Text), re.IGNORECASE)
	return int(Match.group(1)) if Match else None


def IsValidSimpleDecimalNumber(Value: object) -> bool:
	if isinstance(Value, (int, float, decimal.Decimal)):
		return True
	return bool(VALID_SIMPLE_DECIMAL_NUMBER_RE.fullmatch(NormalizeSpaces(Value)))


def ValueToDecimal(Value: object) -> Optional[decimal.Decimal]:
	if Value is None or Value == "":
		return None
	if isinstance(Value, decimal.Decimal):
		return Value
	if isinstance(Value, int):
		return decimal.Decimal(Value)
	if isinstance(Value, float):
		return decimal.Decimal(str(Value))
	Text = NormalizeSpaces(Value)
	if not IsValidSimpleDecimalNumber(Text):
		return None
	return decimal.Decimal(Text.replace(",", "."))


def DecimalToWorkbookNumber(Value: decimal.Decimal) -> float | int:
	if Value == Value.to_integral_value():
		return int(Value)
	return float(Value)


class WorkbookMetadataExtractor:
	def __init__(self, Columns: dgm_database.Columns) -> None:
		self.Columns = Columns

	def Extract(self, FilePath: Path, Workbook: object) -> DgmDocumentMetadata:
		Metadata = DgmDocumentMetadata()
		self._AddSource(Metadata, "filename", ParseDocumentText(FilePath.name))
		Sheet = Workbook.active  # type: ignore[attr-defined]
		Row1 = Sheet["A1"].value
		Row2 = Sheet["A2"].value
		Row1Number = ParseRow1Number(Row1)
		self._AddSource(Metadata, "row1", {"FileNumber": Row1Number} if Row1Number is not None else {})
		self._AddSource(Metadata, "row2", ParseDocumentText(f"Відомість {NormalizeSpaces(Row2)}"))
		Metadata.People = self._FindPeople(Sheet)
		Metadata.FooterPeople = ParseFooterPeople(Metadata.People)
		Metadata.FormularyValues = self._FindFormularyValues(Sheet)
		self._Resolve(Metadata)
		return Metadata

	def _AddSource(self, Metadata: DgmDocumentMetadata, Source: str, Values: Dict[str, object]) -> None:
		Metadata.Sources[Source] = Values

	def _Resolve(self, Metadata: DgmDocumentMetadata) -> None:
		Fields = ("FileNumber", "DocumentType", "EquipmentName", "SerialNumber", "ManufactureYear")
		for Field in Fields:
			Seen: List[Tuple[object, str]] = []
			for Source, Values in Metadata.Sources.items():
				Value = Values.get(Field)
				if Value in (None, ""):
					continue
				if Field == "SerialNumber":
					Value = CanonicalSerial(Value)
				Seen.append((Value, Source))
			if not Seen:
				continue
			Chosen, ChosenSource = next(((Value, Source) for Value, Source in Seen if Source != "filename"), Seen[0])
			setattr(Metadata, Field, Chosen)
			Conflicting = [(Value, Source) for Value, Source in Seen if not self._MetadataValuesCompatible(Field, Chosen, ChosenSource, Value, Source)]
			if Conflicting:
				AllValues: Dict[object, List[str]] = {}
				for Value, Source in Seen:
					AllValues.setdefault(Value, []).append(Source)
				Metadata.Conflicts.append(f"{Field} differs between sources: " + "; ".join(f"{Value} ({', '.join(Sources)})" for Value, Sources in AllValues.items()))

	def _MetadataValuesCompatible(self, Field: str, LeftValue: object, LeftSource: str, RightValue: object, RightSource: str) -> bool:
		if LeftValue == RightValue:
			return True
		if Field in ("EquipmentName", "SerialNumber"):
			if LeftSource == "filename":
				return FilenameCompatibleText(RightValue, LeftValue)
			if RightSource == "filename":
				return FilenameCompatibleText(LeftValue, RightValue)
		return False

	def _FindPeople(self, Sheet: object) -> str:
		for Row in range(1, (Sheet.max_row or 1) + 1):  # type: ignore[attr-defined]
			for Cell in Sheet[Row]:  # type: ignore[index]
				Text = NormalizeSpaces(Cell.value)
				if HEADER_END_MARKER in Text:
					return Text
		return ""

	def _FindFormularyValues(self, Sheet: object) -> Optional[dgm_database.DgmValues]:
		Row = FindLabelRow(Sheet, FORMULARY_LABEL)
		if Row is None:
			return None
		Values = []
		for MetalKey, _ in dgm_database.METALS:
			Value = ValueToDecimal(Sheet[f"{self.Columns.Total[MetalKey]}{Row}"].value)  # type: ignore[index]
			Values.append(Value or decimal.Decimal("0"))
		return dgm_database.DgmValues(*Values)


class XlsxPostprocessor:
	def __init__(self, Database: dgm_database.DgmDatabase, RulesPath: Path) -> None:
		self.Database = Database
		self.Preprocessor = dgm_xlsx_preprocessor.XlsxPreprocessor(Database, RulesPath)
		self.MetadataExtractor = WorkbookMetadataExtractor(Database.Columns)

	def ProcessFile(self, FilePath: Path, ForcedDocumentType: Optional[str] = None, RenameToCanonical: bool = False, FooterStartOverride: Optional[int] = None) -> PostprocessResult:
		if openpyxl is None:
			raise RuntimeError("Missing dependency: openpyxl")
		Workbook = openpyxl.load_workbook(FilePath, data_only=False)
		Sheet = Workbook.active
		Metadata = self.MetadataExtractor.Extract(FilePath, Workbook)
		if ForcedDocumentType in (DOCUMENT_TYPE_PRESENT, DOCUMENT_TYPE_MISSING):
			Metadata.DocumentType = ForcedDocumentType
		if Metadata.DocumentType is None:
			Metadata.DocumentType = self._GuessDocumentType(FilePath, Sheet)
		Issues = self.FindReviewIssues(Sheet)
		HeaderEnd = self._FindHeaderEnd(Sheet) or 6
		FooterStart = FooterStartOverride or self._FindValidFooterStart(Sheet)
		if FooterStart is None:
			raise FooterPlacementRequired(FilePath, self._FindFooterReviewStart(Sheet))
		self._NormalizeHeader(Sheet, Metadata, HeaderEnd)
		self._NormalizeBodyRows(Sheet, HeaderEnd + 1, FooterStart - 1)
		FooterStart = self._RebuildFooter(Sheet, Metadata, FooterStart)
		self._WriteFooterFormulas(Sheet, Metadata, HeaderEnd + 1, FooterStart)
		self._NormalizeFormatting(Sheet)
		self._FitColumnWidths(Sheet)
		Warnings = self._BuildWarnings(Sheet, Metadata, HeaderEnd + 1, FooterStart)
		SavedPath = FilePath
		Workbook.save(FilePath)
		if RenameToCanonical:
			Target = FilePath.with_name(Metadata.CanonicalFilename())
			if Target != FilePath:
				FilePath.rename(Target)
				SavedPath = Target
		return PostprocessResult(FilePath, SavedPath, Metadata, Issues, Warnings)

	def FindReviewIssues(self, Sheet: object) -> List[PostprocessRowIssue]:
		Issues: List[PostprocessRowIssue] = []
		for Row in range(1, (Sheet.max_row or 1) + 1):
			RawName = Sheet[f"{self.Database.Columns.Name}{Row}"].value  # type: ignore[index]
			if not dgm_xlsx_common.CellHasUsableText(RawName):
				continue
			Name = NormalizeSpaces(RawName)
			if self.Database.IsIgnoredText(Name) or self.Preprocessor.IsIgnoredText(Name):
				Issues.append(PostprocessRowIssue(str(Sheet.title), Row, Name, "ignored-text row"))
				continue
			SearchResult = self.Database.FindElement(Name)
			if SearchResult.Record is None or not SearchResult.Record.HasDgm:
				Issues.append(PostprocessRowIssue(str(Sheet.title), Row, Name, "no exact database DGM match"))
			elif not self._HasRowTotalFormulas(Sheet, Row):
				Issues.append(PostprocessRowIssue(str(Sheet.title), Row, Name, "missing row total formulas"))
		return Issues

	def RemoveRows(self, FilePath: Path, Rows: Sequence[int]) -> None:
		Workbook = openpyxl.load_workbook(FilePath, data_only=False)  # type: ignore[union-attr]
		Sheet = Workbook.active
		for Row in sorted(set(Rows), reverse=True):
			Sheet.delete_rows(Row, 1)
		FooterStart = self._FindValidFooterStart(Sheet) or (Sheet.max_row or 1)
		HeaderEnd = self._FindHeaderEnd(Sheet) or 6
		Metadata = self.MetadataExtractor.Extract(FilePath, Workbook)
		self._WriteFooterFormulas(Sheet, Metadata, HeaderEnd + 1, FooterStart)
		self._FitColumnWidths(Sheet)
		Workbook.save(FilePath)

	def SetRowsToZero(self, FilePath: Path, Rows: Sequence[int]) -> None:
		Workbook = openpyxl.load_workbook(FilePath, data_only=False)  # type: ignore[union-attr]
		Sheet = Workbook.active
		for Row in Rows:
			self._ZeroFillRow(Sheet, Row)
		self._WriteFooterFormulas(Sheet, self.MetadataExtractor.Extract(FilePath, Workbook), (self._FindHeaderEnd(Sheet) or 6) + 1, self._FindValidFooterStart(Sheet) or (Sheet.max_row or 1))
		self._FitColumnWidths(Sheet)
		Workbook.save(FilePath)

	def SetRowsInformationMissing(self, FilePath: Path, Rows: Sequence[int]) -> None:
		Workbook = openpyxl.load_workbook(FilePath, data_only=False)  # type: ignore[union-attr]
		Sheet = Workbook.active
		for Row in Rows:
			First = True
			for MetalKey, _ in dgm_database.METALS:
				Sheet[f"{self.Database.Columns.PerElement[MetalKey]}{Row}"].value = INFORMATION_MISSING_TEXT if First else None
				Sheet[f"{self.Database.Columns.Total[MetalKey]}{Row}"].value = None
				First = False
		self._WriteFooterFormulas(Sheet, self.MetadataExtractor.Extract(FilePath, Workbook), (self._FindHeaderEnd(Sheet) or 6) + 1, self._FindValidFooterStart(Sheet) or (Sheet.max_row or 1))
		self._FitColumnWidths(Sheet)
		Workbook.save(FilePath)

	def PutRowFormulas(self, FilePath: Path, Rows: Sequence[int]) -> None:
		Workbook = openpyxl.load_workbook(FilePath, data_only=False)  # type: ignore[union-attr]
		Sheet = Workbook.active
		for Row in Rows:
			self._WriteRowTotalFormulas(Sheet, Row)
		self._WriteFooterFormulas(Sheet, self.MetadataExtractor.Extract(FilePath, Workbook), (self._FindHeaderEnd(Sheet) or 6) + 1, self._FindValidFooterStart(Sheet) or (Sheet.max_row or 1))
		self._FitColumnWidths(Sheet)
		Workbook.save(FilePath)

	def ApplyApprovedMetadata(self, FilePath: Path, Metadata: DgmDocumentMetadata) -> Path:
		Workbook = openpyxl.load_workbook(FilePath, data_only=False)  # type: ignore[union-attr]
		Sheet = Workbook.active
		if Metadata.DocumentType is None:
			Metadata.DocumentType = self._GuessDocumentType(FilePath, Sheet)
		if not Metadata.FooterPeople:
			Metadata.People = self.MetadataExtractor._FindPeople(Sheet)
			Metadata.FooterPeople = ParseFooterPeople(Metadata.People)
		HeaderEnd = self._FindHeaderEnd(Sheet) or 6
		FooterStart = self._FindValidFooterStart(Sheet)
		if FooterStart is None:
			raise FooterPlacementRequired(FilePath, self._FindFooterReviewStart(Sheet))
		self._NormalizeHeader(Sheet, Metadata, HeaderEnd)
		FooterStart = self._RebuildFooter(Sheet, Metadata, FooterStart)
		self._WriteFooterFormulas(Sheet, Metadata, HeaderEnd + 1, FooterStart)
		self._NormalizeFormatting(Sheet)
		self._FitColumnWidths(Sheet)
		Workbook.save(FilePath)
		Target = FilePath.with_name(Metadata.CanonicalFilename())
		if Target != FilePath:
			FilePath.rename(Target)
			return Target
		return FilePath

	def ApplyFormularyValues(self, FilePath: Path, Values: dgm_database.DgmValues) -> None:
		Workbook = openpyxl.load_workbook(FilePath, data_only=False)  # type: ignore[union-attr]
		Sheet = Workbook.active
		Metadata = self.MetadataExtractor.Extract(FilePath, Workbook)
		Metadata.FormularyValues = Values
		HeaderEnd = self._FindHeaderEnd(Sheet) or 6
		FooterStart = self._FindValidFooterStart(Sheet)
		if FooterStart is None:
			raise FooterPlacementRequired(FilePath, self._FindFooterReviewStart(Sheet))
		self._WriteFooterFormularyValues(Sheet, Values)
		self._WriteFooterFormulas(Sheet, Metadata, HeaderEnd + 1, FooterStart)
		self._FitColumnWidths(Sheet)
		Workbook.save(FilePath)

	def CalculateTotalInProduct(self, FilePath: Path) -> dgm_database.DgmValues:
		Workbook = openpyxl.load_workbook(FilePath, data_only=False)  # type: ignore[union-attr]
		Sheet = Workbook.active
		Metadata = self.MetadataExtractor.Extract(FilePath, Workbook)
		HeaderEnd = self._FindHeaderEnd(Sheet) or 6
		FooterStart = self._FindValidFooterStart(Sheet) or (Sheet.max_row or 1)
		BodyTotals = self._CalculateBodyTotals(Sheet, HeaderEnd + 1, FooterStart - 1)
		if Metadata.DocumentType == DOCUMENT_TYPE_MISSING:
			Formulary = Metadata.FormularyValues or dgm_database.DgmValues(decimal.Decimal("0"), decimal.Decimal("0"), decimal.Decimal("0"), decimal.Decimal("0"))
			return dgm_database.DgmValues(*(Formulary.GetMetalValue(MetalKey) - BodyTotals.GetMetalValue(MetalKey) for MetalKey, _ in dgm_database.METALS))
		return BodyTotals

	def _GuessDocumentType(self, FilePath: Path, Sheet: object) -> str:
		TextParts = [FilePath.name]
		for Row in range(1, min(Sheet.max_row or 1, 20) + 1):  # type: ignore[attr-defined]
			for Cell in Sheet[Row]:  # type: ignore[index]
				if Cell.value:
					TextParts.append(str(Cell.value))
		Text = " ".join(TextParts).casefold()
		if "некомплект" in Text or "відсут" in Text:
			return DOCUMENT_TYPE_MISSING
		return DOCUMENT_TYPE_PRESENT

	def _FindHeaderEnd(self, Sheet: object) -> Optional[int]:
		return FindRowContaining(Sheet, HEADER_END_MARKER)

	def _FindValidFooterStart(self, Sheet: object) -> Optional[int]:
		Labels = {TOTAL_IN_PRODUCT_LABEL, FORMULARY_LABEL, MISSING_LABEL}
		for Row in range(1, max((Sheet.max_row or 1) - 1, 1) + 1):  # type: ignore[attr-defined]
			Found = {self._RowFooterLabel(Sheet, Row + Offset) for Offset in range(3)}
			if Found == Labels:
				return Row
		return None

	def _FindFooterReviewStart(self, Sheet: object) -> int:
		Rows = [FindLabelRow(Sheet, Label) for Label in (TOTAL_IN_PRODUCT_LABEL, FORMULARY_LABEL, MISSING_LABEL)]
		Rows = [Row for Row in Rows if Row is not None]
		return min(Rows) if Rows else (Sheet.max_row or 1) + 1  # type: ignore[attr-defined]

	def _RowFooterLabel(self, Sheet: object, Row: int) -> Optional[str]:
		for Cell in Sheet[Row]:  # type: ignore[index]
			Text = NormalizeSpaces(Cell.value)
			if Text in (TOTAL_IN_PRODUCT_LABEL, FORMULARY_LABEL, MISSING_LABEL):
				return Text
		return None

	def _NormalizeHeader(self, Sheet: object, Metadata: DgmDocumentMetadata, HeaderEnd: int) -> None:
		Sheet["A1"].value = f"Відомість № {Metadata.FileNumber}" if Metadata.FileNumber is not None else "Відомість №"
		Sheet["A2"].value = Metadata.CanonicalTitle().replace("Відомість ", "", 1)

	def _RebuildFooter(self, Sheet: object, Metadata: DgmDocumentMetadata, FooterStart: int) -> int:
		if FooterStart < 1:
			FooterStart = 1
		if Sheet.max_row >= FooterStart:  # type: ignore[attr-defined]
			Sheet.delete_rows(FooterStart, (Sheet.max_row or FooterStart) - FooterStart + 1)  # type: ignore[attr-defined]
		Rows = self._FooterRowsTemplate(Metadata)
		for Offset, Values in enumerate(Rows):
			Row = FooterStart + Offset
			for Column, Value in Values.items():
				Sheet[f"{Column}{Row}"].value = Value
		return FooterStart

	def _FooterRowsTemplate(self, Metadata: DgmDocumentMetadata) -> List[Dict[str, str]]:
		DocumentType = Metadata.DocumentType or DOCUMENT_TYPE_PRESENT
		Labels = [MISSING_LABEL, FORMULARY_LABEL, TOTAL_IN_PRODUCT_LABEL] if DocumentType == DOCUMENT_TYPE_MISSING else [TOTAL_IN_PRODUCT_LABEL, FORMULARY_LABEL, MISSING_LABEL]
		Rows = [
			{"D": Labels[0]},
			{"D": Labels[1]},
			{"D": Labels[2]},
			{},
		]
		People = Metadata.FooterPeople or [FooterPerson(CIVILIAN_RANK, "")]
		for Index, Person in enumerate(People):
			Row = {"C": Person.Rank, "E": Person.Name}
			if Index == 0:
				Row["B"] = "Голова комісії: "
			elif Index == 1:
				Row["B"] = "Члени комісії:"
			Rows.append(Row)
		Rows.extend([
			{"B": "__.__.2026"},
			{},
			{"B": "З висновками комісії, щодо комплектності виробу погоджуюсь:"},
			{"B": "Матеріально-відповідальна особа ", "E": "_______________________________________________________________________"},
			{"B": "Начальник кафедри                                ", "E": "_______________________________________________________________________"},
			{"B": "__.__.2026"},
			{},
			{},
		])
		return Rows

	def _NormalizeBodyRows(self, Sheet: object, FirstRow: int, LastRow: int) -> None:
		for Row in range(FirstRow, LastRow + 1):
			if Sheet[f"{self.Database.Columns.Name}{Row}"].value is None:  # type: ignore[index]
				continue
			if self._RowUsesInformationMissing(Sheet, Row):
				continue
			for MetalKey, _ in dgm_database.METALS:
				Cell = Sheet[f"{self.Database.Columns.PerElement[MetalKey]}{Row}"]  # type: ignore[index]
				Value = ValueToDecimal(Cell.value)
				if Value is not None:
					Cell.value = DecimalToWorkbookNumber(Value)
					Cell.number_format = DGM_NUMBER_FORMAT
			self._WriteRowTotalFormulas(Sheet, Row)

	def _ZeroFillRow(self, Sheet: object, Row: int) -> None:
		for MetalKey, _ in dgm_database.METALS:
			Cell = Sheet[f"{self.Database.Columns.PerElement[MetalKey]}{Row}"]
			Value = ValueToDecimal(Cell.value)
			Cell.value = DecimalToWorkbookNumber(Value) if Value is not None else 0
			Cell.number_format = DGM_NUMBER_FORMAT
		self._WriteRowTotalFormulas(Sheet, Row)

	def _WriteRowTotalFormulas(self, Sheet: object, Row: int) -> None:
		if self._RowUsesInformationMissing(Sheet, Row):
			for MetalKey, _ in dgm_database.METALS:
				Sheet[f"{self.Database.Columns.Total[MetalKey]}{Row}"].value = None
			return
		for MetalKey, _ in dgm_database.METALS:
			TotalCell = Sheet[f"{self.Database.Columns.Total[MetalKey]}{Row}"]
			TotalCell.value = f"={self.Database.Columns.Quantity}{Row}*{self.Database.Columns.PerElement[MetalKey]}{Row}"
			TotalCell.number_format = DGM_NUMBER_FORMAT

	def _HasRowTotalFormulas(self, Sheet: object, Row: int) -> bool:
		return all(str(Sheet[f"{self.Database.Columns.Total[MetalKey]}{Row}"].value or "").startswith("=") for MetalKey, _ in dgm_database.METALS)

	def _RowUsesInformationMissing(self, Sheet: object, Row: int) -> bool:
		return any(NormalizeSpaces(Sheet[f"{self.Database.Columns.PerElement[MetalKey]}{Row}"].value) == INFORMATION_MISSING_TEXT for MetalKey, _ in dgm_database.METALS)

	def _WriteFooterFormularyValues(self, Sheet: object, Values: dgm_database.DgmValues, Overwrite: bool = True) -> None:
		FooterRows = self._FooterRows(Sheet)
		if FORMULARY_LABEL not in FooterRows:
			return
		Row = FooterRows[FORMULARY_LABEL]
		for MetalKey, _ in dgm_database.METALS:
			Cell = Sheet[f"{self.Database.Columns.Total[MetalKey]}{Row}"]
			if Overwrite or Cell.value in (None, ""):
				Cell.value = DecimalToWorkbookNumber(Values.GetMetalValue(MetalKey))
			Cell.number_format = DGM_NUMBER_FORMAT

	def _WriteFooterFormulas(self, Sheet: object, Metadata: DgmDocumentMetadata, BodyStart: int, FooterStart: int) -> None:
		FooterRows = self._FooterRows(Sheet)
		if Metadata.FormularyValues is not None:
			self._WriteFooterFormularyValues(Sheet, Metadata.FormularyValues, False)
		BodyEnd = FooterStart - 1
		if BodyEnd < BodyStart:
			BodyEnd = BodyStart
		for MetalKey, _ in dgm_database.METALS:
			TotalColumn = self.Database.Columns.Total[MetalKey]
			if MISSING_LABEL in FooterRows:
				MissingCell = f"{TotalColumn}{FooterRows[MISSING_LABEL]}"
			if TOTAL_IN_PRODUCT_LABEL in FooterRows:
				PresentCell = f"{TotalColumn}{FooterRows[TOTAL_IN_PRODUCT_LABEL]}"
			FormularyCell = f"{TotalColumn}{FooterRows[FORMULARY_LABEL]}" if FORMULARY_LABEL in FooterRows else ""
			if Metadata.DocumentType == DOCUMENT_TYPE_MISSING:
				Sheet[MissingCell].value = f"=SUM({TotalColumn}{BodyStart}:{TotalColumn}{BodyEnd})"
				Sheet[PresentCell].value = f"={FormularyCell}-{MissingCell}" if FormularyCell else ""
			else:
				Sheet[PresentCell].value = f"=SUM({TotalColumn}{BodyStart}:{TotalColumn}{BodyEnd})"
				Sheet[MissingCell].value = f"={FormularyCell}-{PresentCell}" if FormularyCell else ""
			for Row in FooterRows.values():
				Sheet[f"{TotalColumn}{Row}"].number_format = DGM_NUMBER_FORMAT
		self._ApplyDgmBorders(Sheet, BodyStart, FooterStart, FooterRows)

	def _ApplyDgmBorders(self, Sheet: object, BodyStart: int, FooterStart: int, FooterRows: Dict[str, int]) -> None:
		Border = self._DgmCellBorder()
		for Row in range(BodyStart, FooterStart):
			for Column in self._DgmBodyBorderColumns():
				Sheet[f"{Column}{Row}"].border = copy.copy(Border)
		FooterBorderColumns = ["D", *self._DgmBodyBorderColumns()]
		FooterStartColumn = min(openpyxl.utils.column_index_from_string(Column) for Column in FooterBorderColumns)  # type: ignore[union-attr]
		FooterEndColumn = max(openpyxl.utils.column_index_from_string(Column) for Column in FooterBorderColumns)  # type: ignore[union-attr]
		for Row in FooterRows.values():
			for ColumnIndex in range(FooterStartColumn, FooterEndColumn + 1):
				Column = openpyxl.utils.get_column_letter(ColumnIndex)  # type: ignore[union-attr]
				Sheet[f"{Column}{Row}"].border = copy.copy(Border)

	def _DgmBodyBorderColumns(self) -> List[str]:
		Columns = list(self.Database.Columns.PerElement.values()) + list(self.Database.Columns.Total.values())
		return sorted(set(Columns), key=lambda Column: openpyxl.utils.column_index_from_string(Column))  # type: ignore[union-attr]

	@staticmethod
	def _DgmCellBorder() -> object:
		Side = openpyxl.styles.Side(style=DGM_BORDER_STYLE, color=DGM_BORDER_COLOR)  # type: ignore[union-attr]
		return openpyxl.styles.Border(left=Side, right=Side, top=Side, bottom=Side)  # type: ignore[union-attr]

	def _FooterRows(self, Sheet: object) -> Dict[str, int]:
		Rows: Dict[str, int] = {}
		for Label in (TOTAL_IN_PRODUCT_LABEL, FORMULARY_LABEL, MISSING_LABEL):
			Row = FindLabelRow(Sheet, Label)
			if Row is not None:
				Rows[Label] = Row
		return Rows

	def _BuildWarnings(self, Sheet: object, Metadata: DgmDocumentMetadata, BodyStart: int, FooterStart: int) -> List[str]:
		Warnings: List[str] = []
		FooterRows = self._FooterRows(Sheet)
		Formulary = Metadata.FormularyValues or self.MetadataExtractor._FindFormularyValues(Sheet)
		if Formulary is None:
			Warnings.append("Formulary DGM values are missing.")
			return Warnings
		BodyTotals = self._CalculateBodyTotals(Sheet, BodyStart, FooterStart - 1)
		for MetalKey, MetalName in dgm_database.METALS:
			Value = Formulary.GetMetalValue(MetalKey) - BodyTotals.GetMetalValue(MetalKey)
			if Value < 0:
				Warnings.append(f"{MetalName} calculated footer value is negative: {dgm_database.DecimalToText(Value)}")
		return Warnings

	def _CalculateBodyTotals(self, Sheet: object, FirstRow: int, LastRow: int) -> dgm_database.DgmValues:
		Values = dgm_database.DgmValues(decimal.Decimal("0"), decimal.Decimal("0"), decimal.Decimal("0"), decimal.Decimal("0"))
		for Row in range(FirstRow, LastRow + 1):
			if self._RowUsesInformationMissing(Sheet, Row):
				continue
			Qty = ValueToDecimal(Sheet[f"{self.Database.Columns.Quantity}{Row}"].value) or decimal.Decimal("0")
			for MetalKey, _ in dgm_database.METALS:
				Per = ValueToDecimal(Sheet[f"{self.Database.Columns.PerElement[MetalKey]}{Row}"].value) or decimal.Decimal("0")
				Values.SetMetalValue(MetalKey, Values.GetMetalValue(MetalKey) + Qty * Per)
		return Values

	def _NormalizeFormatting(self, Sheet: object) -> None:
		Black = "000000"
		for Row in Sheet.iter_rows():
			for Cell in Row:
				if isinstance(Cell, MergedCell):
					continue
				Cell.fill = copy.copy(openpyxl.styles.PatternFill(fill_type=None))  # type: ignore[union-attr]
				Font = copy.copy(Cell.font)
				Font.color = Black
				if Cell.column_letter in set(self.Database.Columns.PerElement.values()) | set(self.Database.Columns.Total.values()):
					Font.bold = False
					Font.underline = None
				Cell.font = Font
				if Cell.column_letter in set(self.Database.Columns.PerElement.values()) | set(self.Database.Columns.Total.values()):
					Cell.number_format = DGM_NUMBER_FORMAT if Cell.value != INFORMATION_MISSING_TEXT else "@"

	def _FitColumnWidths(self, Sheet: object) -> None:
		MergedCoordinates = {Cell.coordinate for Range in Sheet.merged_cells.ranges for Row in Sheet[Range.coord] for Cell in Row}
		for ColumnIndex in range(1, (Sheet.max_column or 1) + 1):
			Letter = openpyxl.utils.get_column_letter(ColumnIndex)  # type: ignore[union-attr]
			MaxWidth = MIN_COLUMN_WIDTH
			for Row in range(1, (Sheet.max_row or 1) + 1):
				Cell = Sheet.cell(Row, ColumnIndex)
				if Cell.coordinate in MergedCoordinates or isinstance(Cell, MergedCell):
					continue
				Value = Cell.value
				if Value in (None, ""):
					continue
				IsNumeric = isinstance(Value, (int, float, decimal.Decimal)) or str(Value).startswith("=")
				RightValue = Sheet.cell(Row, ColumnIndex + 1).value if ColumnIndex < (Sheet.max_column or 1) else "occupied"
				if not IsNumeric and RightValue in (None, ""):
					continue
				Width = len(str(Value)) + WIDTH_PADDING
				MaxWidth = max(MaxWidth, Width)
			Limit = MAX_TEXT_COLUMN_WIDTH if Letter == self.Database.Columns.Name else MAX_NUMERIC_COLUMN_WIDTH
			Sheet.column_dimensions[Letter].width = min(max(MaxWidth, MIN_COLUMN_WIDTH), Limit)


def FindRowContaining(Sheet: object, Needle: str) -> Optional[int]:
	for Row in range(1, (Sheet.max_row or 1) + 1):  # type: ignore[attr-defined]
		for Cell in Sheet[Row]:  # type: ignore[index]
			if Needle in NormalizeSpaces(Cell.value):
				return Row
	return None


def FindLabelRow(Sheet: object, Label: str) -> Optional[int]:
	for Row in range(1, (Sheet.max_row or 1) + 1):  # type: ignore[attr-defined]
		for Cell in Sheet[Row]:  # type: ignore[index]
			if NormalizeSpaces(Cell.value) == Label:
				return Row
	return None
