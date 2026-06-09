#!/usr/bin/env python3
"""
Fill DGM metal content columns in XLSX inventory files using a small XML database.

Required dependency:
    pip install openpyxl

Usage:
    python dgm_inventory.py database.xml folder_with_xlsx_files
"""

from __future__ import annotations

import copy
import decimal
import shutil
import sys
import xml.etree.ElementTree as XmlTree
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Set, Tuple

try:
	import openpyxl
	import openpyxl.utils
except ImportError:
	print("Missing dependency: openpyxl", file=sys.stderr)
	print("Install it with: pip install openpyxl", file=sys.stderr)
	sys.exit(2)


CREATE_BACKUP = True
STOP_AFTER_CONSECUTIVE_IGNORED_ROWS = 20
DEFAULT_SHEET_MODE = "active"  # This script processes only the active sheet in each workbook.

METALS: Tuple[Tuple[str, str], ...] = (
	("gold", "Gold"),
	("silver", "Silver"),
	("platinum", "Platinum"),
	("mpg", "MPG/palladium"),
)

DEFAULT_COLUMNS = {
	"name": "B",
	"quantity": "C",
	"gold": "D",
	"silver": "E",
	"platinum": "F",
	"mpg": "G",
	"total_gold": "H",
	"total_silver": "I",
	"total_platinum": "J",
	"total_mpg": "K",
}

GRAM_NUMBER_FORMAT = "0.000000"


@dataclass
class Columns:
	Name: str
	Quantity: str
	PerElement: Dict[str, str]
	Total: Dict[str, str]


@dataclass
class DgmEntry:
	Name: str
	GoldG: decimal.Decimal
	SilverG: decimal.Decimal
	PlatinumG: decimal.Decimal
	MpgG: decimal.Decimal

	def GetMetalValue(self, MetalKey: str) -> decimal.Decimal:
		if MetalKey == "gold":
			return self.GoldG
		if MetalKey == "silver":
			return self.SilverG
		if MetalKey == "platinum":
			return self.PlatinumG
		if MetalKey == "mpg":
			return self.MpgG
		raise ValueError(f"Unknown metal key: {MetalKey}")


class Database:
	def __init__(self, PathToDatabase: Path) -> None:
		self.Path = PathToDatabase
		self.Tree: XmlTree.ElementTree
		self.Root: XmlTree.Element
		self.SettingsNode: XmlTree.Element
		self.ColumnsNode: XmlTree.Element
		self.ElementsNode: XmlTree.Element
		self.IgnoredNode: XmlTree.Element
		self.Columns: Columns
		self.Elements: Dict[str, DgmEntry] = {}
		self.IgnoredTexts: Set[str] = set()

		if self.Path.exists():
			self.Load()
		else:
			self.CreateEmpty()
			self.Save()

	def CreateEmpty(self) -> None:
		self.Root = XmlTree.Element("dgm_database", {"version": "1"})
		self.SettingsNode = XmlTree.SubElement(self.Root, "settings")
		self.ColumnsNode = XmlTree.SubElement(self.SettingsNode, "columns", copy.deepcopy(DEFAULT_COLUMNS))
		self.ElementsNode = XmlTree.SubElement(self.Root, "elements")
		self.IgnoredNode = XmlTree.SubElement(self.Root, "ignored")
		self.Tree = XmlTree.ElementTree(self.Root)
		self.Columns = self.ReadColumns()

	def Load(self) -> None:
		try:
			self.Tree = XmlTree.parse(self.Path)
		except XmlTree.ParseError as Error:
			raise RuntimeError(f"Cannot parse XML database '{self.Path}': {Error}") from Error

		self.Root = self.Tree.getroot()
		if self.Root.tag != "dgm_database":
			raise RuntimeError(f"Unexpected XML root in '{self.Path}': {self.Root.tag}")

		self.SettingsNode = self.EnsureChild(self.Root, "settings")
		self.ColumnsNode = self.EnsureChild(self.SettingsNode, "columns", DEFAULT_COLUMNS)
		self.ElementsNode = self.EnsureChild(self.Root, "elements")
		self.IgnoredNode = self.EnsureChild(self.Root, "ignored")
		self.Columns = self.ReadColumns()
		self.LoadElements()
		self.LoadIgnoredTexts()

	def EnsureChild(self, Parent: XmlTree.Element, Tag: str, Attributes: Optional[Dict[str, str]] = None) -> XmlTree.Element:
		Existing = Parent.find(Tag)
		if Existing is not None:
			if Attributes is not None:
				for Key, Value in Attributes.items():
					Existing.set(Key, Existing.get(Key, Value))
			return Existing
		return XmlTree.SubElement(Parent, Tag, Attributes or {})

	def ReadColumns(self) -> Columns:
		def ReadColumn(Attribute: str) -> str:
			Column = self.ColumnsNode.get(Attribute, DEFAULT_COLUMNS[Attribute]).strip().upper()
			try:
				openpyxl.utils.column_index_from_string(Column)
			except ValueError as Error:
				raise RuntimeError(f"Invalid column '{Column}' in database setting '{Attribute}'") from Error
			return Column

		PerElement = {}
		Total = {}
		for MetalKey, _ in METALS:
			PerElement[MetalKey] = ReadColumn(MetalKey)
			Total[MetalKey] = ReadColumn(f"total_{MetalKey}")

		return Columns(
			Name=ReadColumn("name"),
			Quantity=ReadColumn("quantity"),
			PerElement=PerElement,
			Total=Total,
		)

	def LoadElements(self) -> None:
		self.Elements.clear()
		for Node in self.ElementsNode.findall("element"):
			Name = Node.get("name", "").strip()
			if not Name:
				continue
			Key = Node.get("key") or NormalizeText(Name)
			Entry = DgmEntry(
				Name=Name,
				GoldG=ReadDecimal(Node.get("gold_g", "0")),
				SilverG=ReadDecimal(Node.get("silver_g", "0")),
				PlatinumG=ReadDecimal(Node.get("platinum_g", "0")),
				MpgG=ReadDecimal(Node.get("mpg_g", "0")),
			)
			self.Elements[Key] = Entry

	def LoadIgnoredTexts(self) -> None:
		self.IgnoredTexts.clear()
		for Node in self.IgnoredNode.findall("text"):
			Value = Node.get("value", "")
			Key = Node.get("key") or NormalizeText(Value)
			if Key:
				self.IgnoredTexts.add(Key)

	def AddElement(self, Name: str, Entry: DgmEntry) -> None:
		Key = NormalizeText(Name)
		if Key in self.Elements:
			self.Elements[Key] = Entry
			return

		XmlTree.SubElement(
			self.ElementsNode,
			"element",
			{
				"key": Key,
				"name": Name,
				"gold_g": DecimalToText(Entry.GoldG),
				"silver_g": DecimalToText(Entry.SilverG),
				"platinum_g": DecimalToText(Entry.PlatinumG),
				"mpg_g": DecimalToText(Entry.MpgG),
			},
		)
		self.Elements[Key] = Entry

	def AddIgnoredText(self, Text: str) -> None:
		Key = NormalizeText(Text)
		if not Key or Key in self.IgnoredTexts:
			return
		XmlTree.SubElement(self.IgnoredNode, "text", {"key": Key, "value": Text})
		self.IgnoredTexts.add(Key)

	def Save(self) -> None:
		self.Path.parent.mkdir(parents=True, exist_ok=True)
		try:
			XmlTree.indent(self.Tree, space="\t", level=0)
		except AttributeError:
			pass
		self.Tree.write(self.Path, encoding="utf-8", xml_declaration=True)


def NormalizeText(Value: str) -> str:
	return " ".join(Value.strip().split()).casefold()


def ReadDecimal(Value: str) -> decimal.Decimal:
	Value = (Value or "0").strip().replace(",", ".")
	if Value == "":
		Value = "0"
	return decimal.Decimal(Value)


def DecimalToText(Value: decimal.Decimal) -> str:
	if Value == 0:
		return "0"
	return format(Value.normalize(), "f")


def CellHasUsableText(Value: object) -> bool:
	return isinstance(Value, str) and bool(Value.strip())


def DecimalToExcelNumber(Value: decimal.Decimal) -> float:
	return float(Value)


def AskElementOrIgnore(FilePath: Path, SheetName: str, Row: int, Text: str) -> Optional[DgmEntry]:
	print("\nUnknown text found:")
	print(f"  File:  {FilePath}")
	print(f"  Sheet: {SheetName}")
	print(f"  Row:   {Row}")
	print(f"  Text:  {Text}")

	while True:
		Answer = input("Is this a radio element? [e] element / [i] ignore: ").strip().casefold()
		if Answer in ("e", "element"):
			return AskDgmValues(Text)
		if Answer in ("i", "ignore"):
			return None
		print("Please enter 'e' or 'i'.")


def AskDgmValues(Name: str) -> DgmEntry:
	print(f"Enter DGM content for: {Name}")
	print("Values are entered in milligrams. Empty input means 0 mg.")

	GoldMg = AskNonNegativeDecimal("Gold, mg")
	SilverMg = AskNonNegativeDecimal("Silver, mg")
	PlatinumMg = AskNonNegativeDecimal("Platinum, mg")
	MpgMg = AskNonNegativeDecimal("MPG/palladium, mg")

	return DgmEntry(
		Name=Name,
		GoldG=GoldMg / decimal.Decimal("1000"),
		SilverG=SilverMg / decimal.Decimal("1000"),
		PlatinumG=PlatinumMg / decimal.Decimal("1000"),
		MpgG=MpgMg / decimal.Decimal("1000"),
	)


def AskNonNegativeDecimal(Prompt: str) -> decimal.Decimal:
	while True:
		Raw = input(f"  {Prompt} [0]: ").strip().replace(",", ".")
		if Raw == "":
			return decimal.Decimal("0")
		try:
			Value = decimal.Decimal(Raw)
		except decimal.InvalidOperation:
			print("Please enter a number, for example: 12.5")
			continue
		if Value < 0:
			print("The value cannot be negative.")
			continue
		return Value


def WriteEntryToRow(Sheet: openpyxl.worksheet.worksheet.Worksheet, Row: int, ColumnsInfo: Columns, Entry: DgmEntry) -> None:
	for MetalKey, _ in METALS:
		PerElementCell = Sheet[f"{ColumnsInfo.PerElement[MetalKey]}{Row}"]
		TotalCell = Sheet[f"{ColumnsInfo.Total[MetalKey]}{Row}"]
		PerElementCell.value = DecimalToExcelNumber(Entry.GetMetalValue(MetalKey))
		PerElementCell.number_format = GRAM_NUMBER_FORMAT
		TotalCell.value = f"={ColumnsInfo.PerElement[MetalKey]}{Row}*{ColumnsInfo.Quantity}{Row}"
		TotalCell.number_format = GRAM_NUMBER_FORMAT


def WriteWorkbookTotals(Sheet: openpyxl.worksheet.worksheet.Worksheet, TotalRow: int, ColumnsInfo: Columns, ProcessedRows: List[int]) -> None:
	for MetalKey, _ in METALS:
		TotalColumn = ColumnsInfo.Total[MetalKey]
		Cell = Sheet[f"{TotalColumn}{TotalRow}"]
		Cell.value = BuildSumFormula(TotalColumn, ProcessedRows)
		Cell.number_format = GRAM_NUMBER_FORMAT


def BuildSumFormula(Column: str, Rows: Iterable[int]) -> str:
	SortedRows = sorted(set(Rows))
	if not SortedRows:
		return "=0"

	Ranges: List[str] = []
	Start = SortedRows[0]
	Previous = SortedRows[0]

	for Row in SortedRows[1:]:
		if Row == Previous + 1:
			Previous = Row
			continue
		Ranges.append(FormatRange(Column, Start, Previous))
		Start = Row
		Previous = Row

	Ranges.append(FormatRange(Column, Start, Previous))
	return f"=SUM({','.join(Ranges)})"


def FormatRange(Column: str, Start: int, End: int) -> str:
	if Start == End:
		return f"{Column}{Start}"
	return f"{Column}{Start}:{Column}{End}"


def ProcessWorkbook(FilePath: Path, Db: Database) -> Tuple[int, int]:
	Workbook = openpyxl.load_workbook(FilePath, data_only=False)
	Sheet = Workbook.active

	ProcessedRows: List[int] = []
	IgnoredRows: Set[int] = set()
	LastProcessedRow = 0
	ConsecutiveIgnoredRows = 0
	Row = 1
	MaxRow = Sheet.max_row or 1

	while Row <= MaxRow:
		RawName = Sheet[f"{Db.Columns.Name}{Row}"].value

		if not CellHasUsableText(RawName):
			IgnoredRows.add(Row)
			ConsecutiveIgnoredRows += 1
		else:
			Name = " ".join(str(RawName).strip().split())
			Key = NormalizeText(Name)

			if Key in Db.IgnoredTexts:
				IgnoredRows.add(Row)
				ConsecutiveIgnoredRows += 1
			else:
				Entry = Db.Elements.get(Key)
				if Entry is None:
					Entry = AskElementOrIgnore(FilePath, Sheet.title, Row, Name)
					if Entry is None:
						Db.AddIgnoredText(Name)
						Db.Save()
						IgnoredRows.add(Row)
						ConsecutiveIgnoredRows += 1
					else:
						Db.AddElement(Name, Entry)
						Db.Save()

				if Entry is not None:
					WriteEntryToRow(Sheet, Row, Db.Columns, Entry)
					ProcessedRows.append(Row)
					LastProcessedRow = Row
					ConsecutiveIgnoredRows = 0

		if ConsecutiveIgnoredRows >= STOP_AFTER_CONSECUTIVE_IGNORED_ROWS:
			break

		Row += 1

	TotalRow = LastProcessedRow + 1 if LastProcessedRow > 0 else 1
	WriteWorkbookTotals(Sheet, TotalRow, Db.Columns, ProcessedRows)

	if CREATE_BACKUP:
		BackupPath = FilePath.with_name(FilePath.name + ".bak")
		if not BackupPath.exists():
			shutil.copy2(FilePath, BackupPath)

	Workbook.save(FilePath)
	return len(ProcessedRows), len(IgnoredRows)


def FindXlsxFiles(Folder: Path) -> List[Path]:
	return sorted(
		PathItem
		for PathItem in Folder.iterdir()
		if PathItem.is_file()
		and PathItem.suffix.casefold() == ".xlsx"
		and not PathItem.name.startswith("~$")
	)


def Main() -> int:
	if len(sys.argv) != 3:
		print("Usage: python dgm_inventory.py database.xml folder_with_xlsx_files", file=sys.stderr)
		return 2

	DatabasePath = Path(sys.argv[1]).expanduser().resolve()
	Folder = Path(sys.argv[2]).expanduser().resolve()

	if not Folder.exists() or not Folder.is_dir():
		print(f"Folder does not exist or is not a directory: {Folder}", file=sys.stderr)
		return 2

	try:
		Db = Database(DatabasePath)
	except Exception as Error:
		print(str(Error), file=sys.stderr)
		return 1

	Files = FindXlsxFiles(Folder)
	if not Files:
		print(f"No .xlsx files found in: {Folder}")
		return 0

	print(f"Database: {DatabasePath}")
	print(f"Folder:   {Folder}")
	print(f"Files:    {len(Files)}")
	print(f"Sheet mode: {DEFAULT_SHEET_MODE}")

	TotalProcessed = 0
	TotalIgnored = 0

	for FilePath in Files:
		print(f"\nProcessing: {FilePath.name}")
		try:
			Processed, Ignored = ProcessWorkbook(FilePath, Db)
		except Exception as Error:
			print(f"ERROR while processing '{FilePath}': {Error}", file=sys.stderr)
			continue

		TotalProcessed += Processed
		TotalIgnored += Ignored
		print(f"  Processed element rows: {Processed}")
		print(f"  Ignored rows:           {Ignored}")

	print("\nDone.")
	print(f"Total processed element rows: {TotalProcessed}")
	print(f"Total ignored rows:           {TotalIgnored}")
	return 0


if __name__ == "__main__":
	sys.exit(Main())
