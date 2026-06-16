#!/usr/bin/env python3
"""
GUI editor for DGM XML databases.

Scope:
- visualize and edit the structured component catalog in a tree view;
- move catalog nodes between parents;
- visualize ignored item texts in a list view.

Usage:
    python dgm_gui.py [database.xml]

If the database argument is omitted, a file selection dialog is shown.
"""

from __future__ import annotations

import decimal
import sys
import tkinter as tk
import tkinter.filedialog
import tkinter.messagebox
import tkinter.ttk as ttk
import xml.etree.ElementTree as XmlTree
from pathlib import Path
from dataclasses import dataclass
from typing import Dict, Optional, List, Tuple

import dgm_database
import dgm_inventory

try:
	import openpyxl
except ImportError:
	openpyxl = None  # type: ignore[assignment]


@dataclass
class GuiMissingElement:
	FilePath: Path
	SheetName: str
	Row: int
	Name: str


@dataclass
class GuiConflictRow:
	FilePath: Path
	SheetName: str
	Row: int
	Name: str
	Record: dgm_database.ElementRecord
	SheetValues: dgm_database.DgmValues
	DatabaseValues: dgm_database.DgmValues
	Details: str


@dataclass
class GuiProcessResult:
	FilePath: Path
	ProcessedRows: int
	IgnoredRows: int
	Missing: List[GuiMissingElement]
	Conflicts: List[GuiConflictRow]


WINDOW_TITLE = "DGM Database Editor"


class DgmDatabaseViewer(tk.Tk):
	def __init__(self, DatabasePath: Path) -> None:
		super().__init__()
		self.DatabasePath = DatabasePath
		self.Database = dgm_database.OpenDatabase(DatabasePath)
		self.CatalogItems: Dict[str, XmlTree.Element] = {}

		self.title(f"{WINDOW_TITLE} - {DatabasePath.name}")
		self.geometry("1100x700")
		self.minsize(850, 500)

		self._ConfigureStyle()
		self._BuildLayout()
		self._PopulateDatabaseViews()

	def _ConfigureStyle(self) -> None:
		Style = ttk.Style(self)
		if "clam" in Style.theme_names():
			Style.theme_use("clam")
		Style.configure("Treeview", rowheight=24)
		Style.configure("Heading.TLabel", font=("TkDefaultFont", 10, "bold"))

	def _BuildLayout(self) -> None:
		self.columnconfigure(0, weight=1)
		self.rowconfigure(1, weight=1)

		Header = ttk.Frame(self, padding=(10, 8, 10, 4))
		Header.grid(row=0, column=0, sticky="ew")
		Header.columnconfigure(1, weight=1)

		ttk.Label(Header, text="Database:", style="Heading.TLabel").grid(row=0, column=0, sticky="w")
		ttk.Label(Header, text=str(self.DatabasePath)).grid(row=0, column=1, sticky="w", padx=(6, 0))
		ButtonFrame = ttk.Frame(Header)
		ButtonFrame.grid(row=0, column=2, sticky="e")
		ttk.Button(ButtonFrame, text="Fill XLSX file...", command=self._SelectAndProcessXlsxFile).grid(row=0, column=0, padx=(0, 6))
		ttk.Button(ButtonFrame, text="Fill XLSX folder...", command=self._SelectAndProcessXlsxFolder).grid(row=0, column=1)

		Content = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
		Content.grid(row=1, column=0, sticky="nsew", padx=10, pady=(4, 10))

		CatalogPane = ttk.Frame(Content)
		IgnoredPane = ttk.Frame(Content)
		Content.add(CatalogPane, weight=4)
		Content.add(IgnoredPane, weight=1)

		self._BuildCatalogPane(CatalogPane)
		self._BuildIgnoredPane(IgnoredPane)

	def _BuildCatalogPane(self, Parent: ttk.Frame) -> None:
		Parent.columnconfigure(0, weight=1)
		Parent.rowconfigure(1, weight=1)

		HeaderFrame = ttk.Frame(Parent)
		HeaderFrame.grid(row=0, column=0, sticky="ew", pady=(0, 4))
		HeaderFrame.columnconfigure(1, weight=1)

		ttk.Label(HeaderFrame, text="Component database", style="Heading.TLabel").grid(row=0, column=0, sticky="w", padx=(0, 10))
		self.SearchEntry = ttk.Entry(HeaderFrame)
		self.SearchEntry.grid(row=0, column=1, sticky="ew", padx=(0, 6))
		ttk.Button(HeaderFrame, text="Search", command=self._SearchCatalog).grid(row=0, column=2)
		self.SearchStatusLabel = ttk.Label(Parent, text="Enter a component name and click Search to test exact and partial matches.")
		self.SearchStatusLabel.grid(row=3, column=0, sticky="ew", pady=(4, 0))
		self.SearchEntry.bind("<Return>", lambda _Event: self._SearchCatalog())

		Columns = ("kind", "name", "gold", "silver", "platinum", "mpg", "pattern")
		self.CatalogTree = ttk.Treeview(Parent, columns=Columns, show="tree headings", selectmode="browse")
		self.CatalogTree.heading("#0", text="Component / structure")
		self.CatalogTree.heading("kind", text="Type")
		self.CatalogTree.heading("name", text="Name")
		self.CatalogTree.heading("gold", text="Gold g")
		self.CatalogTree.heading("silver", text="Silver g")
		self.CatalogTree.heading("platinum", text="Platinum g")
		self.CatalogTree.heading("mpg", text="MPG g")
		self.CatalogTree.heading("pattern", text="Regex pattern")

		self.CatalogTree.column("#0", width=260, minwidth=180, stretch=True)
		self.CatalogTree.column("kind", width=85, minwidth=65, anchor="center", stretch=False)
		self.CatalogTree.column("name", width=180, minwidth=120, stretch=True)
		for Column in ("gold", "silver", "platinum", "mpg"):
			self.CatalogTree.column(Column, width=90, minwidth=70, anchor="e", stretch=False)
		self.CatalogTree.column("pattern", width=180, minwidth=120, stretch=True)

		VerticalScrollbar = ttk.Scrollbar(Parent, orient=tk.VERTICAL, command=self.CatalogTree.yview)
		HorizontalScrollbar = ttk.Scrollbar(Parent, orient=tk.HORIZONTAL, command=self.CatalogTree.xview)
		self.CatalogTree.configure(yscrollcommand=VerticalScrollbar.set, xscrollcommand=HorizontalScrollbar.set)

		self.CatalogTree.grid(row=1, column=0, sticky="nsew")
		VerticalScrollbar.grid(row=1, column=1, sticky="ns")
		HorizontalScrollbar.grid(row=2, column=0, sticky="ew")

		self.CatalogContextMenu = tk.Menu(self, tearoff=False)
		self.CatalogContextMenu.add_command(label="Modify database node...", command=self._ModifySelectedCatalogNode)
		self.CatalogContextMenu.add_command(label="Move node to another parent...", command=self._MoveSelectedCatalogNode)
		self.CatalogTree.bind("<Button-3>", self._ShowCatalogContextMenu)
		self.CatalogTree.bind("<Control-Button-1>", self._ShowCatalogContextMenu)

	def _BuildIgnoredPane(self, Parent: ttk.Frame) -> None:
		Parent.columnconfigure(0, weight=1)
		Parent.rowconfigure(1, weight=1)

		ttk.Label(Parent, text="Ignored items", style="Heading.TLabel").grid(row=0, column=0, sticky="w", pady=(0, 4))

		Frame = ttk.Frame(Parent)
		Frame.grid(row=1, column=0, sticky="nsew")
		Frame.columnconfigure(0, weight=1)
		Frame.rowconfigure(0, weight=1)

		self.IgnoredList = tk.Listbox(Frame, activestyle="dotbox", exportselection=False)
		VerticalScrollbar = ttk.Scrollbar(Frame, orient=tk.VERTICAL, command=self.IgnoredList.yview)
		HorizontalScrollbar = ttk.Scrollbar(Frame, orient=tk.HORIZONTAL, command=self.IgnoredList.xview)
		self.IgnoredList.configure(yscrollcommand=VerticalScrollbar.set, xscrollcommand=HorizontalScrollbar.set)

		self.IgnoredList.grid(row=0, column=0, sticky="nsew")
		VerticalScrollbar.grid(row=0, column=1, sticky="ns")
		HorizontalScrollbar.grid(row=1, column=0, sticky="ew")

		self.IgnoredCountLabel = ttk.Label(Parent, text="0 ignored items")
		self.IgnoredCountLabel.grid(row=2, column=0, sticky="w", pady=(4, 0))

	def _PopulateDatabaseViews(self) -> None:
		self._PopulateCatalogTree()
		self._PopulateIgnoredList()

	def _PopulateCatalogTree(self) -> None:
		self.CatalogItems.clear()
		for Item in self.CatalogTree.get_children():
			self.CatalogTree.delete(Item)

		CatalogRoot = self.CatalogTree.insert("", "end", text="catalog", values=("root", "", "", "", "", "", ""), open=True)
		for Child in list(self.Database.CatalogNode):
			self._InsertCatalogNode(CatalogRoot, Child)

		LegacyElements = self.Database.LegacyElementsNode.findall("element")
		if LegacyElements:
			LegacyRoot = self.CatalogTree.insert("", "end", text="legacy elements", values=("legacy", "", "", "", "", "", ""), open=True)
			for ElementNode in LegacyElements:
				self._InsertLegacyElement(LegacyRoot, ElementNode)

		self._SortCatalogTreeChildren()

	def _SortCatalogTreeChildren(self, ParentItemId: str = "") -> None:
		Children = list(self.CatalogTree.get_children(ParentItemId))

		Children.sort(key=self._CatalogTreeSortKey)

		for Index, ChildItemId in enumerate(Children):
			self.CatalogTree.move(ChildItemId, ParentItemId, Index)
			self._SortCatalogTreeChildren(ChildItemId)

	def _CatalogTreeSortKey(self, ItemId: str) -> tuple[str]:
		Text = self.CatalogTree.item(ItemId, "text")
		return (Text.casefold(),)

	def _InsertCatalogNode(self, ParentItem: str, Node: XmlTree.Element) -> None:
		if Node.tag not in ("node", "regex"):
			return

		Text = '|' + Node.get("text", Node.get("name", Node.tag)) + '|'
		Item = self.CatalogTree.insert(
			ParentItem,
			"end",
			text=Text,
			values=(
				"regex" if Node.tag == "regex" else "node",
				Node.get("name", ""),
				*self._ReadDgmColumns(Node),
				Node.get("pattern", ""),
			),
			open=False,
		)
		self.CatalogItems[Item] = Node
		for Child in list(Node):
			self._InsertCatalogNode(Item, Child)

	def _InsertLegacyElement(self, ParentItem: str, Node: XmlTree.Element) -> None:
		Name = Node.get("name", "")
		self.CatalogTree.insert(
			ParentItem,
			"end",
			text=Name,
			values=("legacy", Name, *self._ReadDgmColumns(Node), ""),
		)

	def _ReadDgmColumns(self, Node: XmlTree.Element) -> tuple[str, str, str, str]:
		DgmNode = Node.find("dgm")
		SourceNode = DgmNode if DgmNode is not None else Node
		if not self._HasDgmValues(SourceNode):
			return ("", "", "", "")
		return tuple(SourceNode.get(f"{MetalKey}_g", "0") for MetalKey, _ in dgm_database.METALS)  # type: ignore[return-value]

	def _HasDgmValues(self, Node: XmlTree.Element) -> bool:
		return any(Node.get(f"{MetalKey}_g") is not None for MetalKey, _ in dgm_database.METALS)


	def _ShowCatalogContextMenu(self, Event: tk.Event) -> str:
		ItemId = self.CatalogTree.identify_row(Event.y)
		if ItemId == "" or ItemId not in self.CatalogItems:
			return "break"

		self.CatalogTree.selection_set(ItemId)
		self.CatalogTree.focus(ItemId)
		self.CatalogContextMenu.tk_popup(Event.x_root, Event.y_root)
		return "break"

	def _GetSelectedCatalogNode(self) -> Optional[XmlTree.Element]:
		ItemId = self.CatalogTree.focus()
		if not ItemId:
			Selection = self.CatalogTree.selection()
			ItemId = Selection[0] if Selection else ""
		return self.CatalogItems.get(ItemId)

	def _ModifySelectedCatalogNode(self) -> None:
		Node = self._GetSelectedCatalogNode()
		if Node is None:
			return

		Dialog = CatalogNodeEditDialog(self, Node)
		if Dialog.Result is None:
			return

		try:
			self._ApplyCatalogNodeChanges(Node, Dialog.Result)
			self.Database.Save()
		except Exception as Error:
			tkinter.messagebox.showerror(WINDOW_TITLE, str(Error), parent=self)
			return

		self._PopulateCatalogTree()

	def _ApplyCatalogNodeChanges(self, Node: XmlTree.Element, Values: Dict[str, str]) -> None:
		Text = Values["text"]
		Name = Values["name"]
		if not Text:
			raise RuntimeError("Display text is required")

		Node.set("text", Text)
		if Name:
			Node.set("name", Name)
		elif "name" in Node.attrib:
			del Node.attrib["name"]

		if Node.tag == "regex":
			Pattern = Values["pattern"]
			if not Pattern:
				raise RuntimeError("Regex pattern is required for regex nodes")
			Node.set("pattern", Pattern)

		DgmNode = self.Database.EnsureChild(Node, "dgm")
		for MetalKey, _ in dgm_database.METALS:
			Value = dgm_database.ReadDecimal(Values[MetalKey])
			DgmNode.set(f"{MetalKey}_g", dgm_database.DecimalToText(Value))

	def _MoveSelectedCatalogNode(self) -> None:
		Node = self._GetSelectedCatalogNode()
		if Node is None:
			return

		Dialog = MoveCatalogNodeDialog(self, self.Database.GetNodePathParts(Node)[:-1])
		if Dialog.Result is None:
			return

		try:
			self.Database.MoveCatalogNode(Node, Dialog.Result)
			self.Database.Save()
		except Exception as Error:
			tkinter.messagebox.showerror(WINDOW_TITLE, str(Error), parent=self)
			return

		self._PopulateCatalogTree()


	def _SearchCatalog(self) -> None:
		SearchText = self.SearchEntry.get().strip()
		if not SearchText:
			self.SearchStatusLabel.configure(text="Enter a component name to search.")
			return

		Result = self.Database.FindElement(SearchText)
		Messages: List[str] = []
		if Result.Record is not None:
			MatchType = "regex" if Result.MatchedByRegex else "exact"
			Messages.append(f"Found {MatchType} match: {Result.Record.DisplayName}")
		else:
			Messages.append("No exact match found.")

		PartialMatches = Result.PartialMatches or []
		if PartialMatches:
			PartialText = "; ".join(
				f"{Match.DisplayName}{' (has DGM)' if Match.HasDgm else ''}"
				for Match in PartialMatches
			)
			Messages.append(f"Partial matches: {PartialText}")
		else:
			Messages.append("No partial matches.")

		self.SearchStatusLabel.configure(text=" ".join(Messages))

	def _SelectAndProcessXlsxFile(self) -> None:
		if openpyxl is None:
			tkinter.messagebox.showerror(WINDOW_TITLE, "Missing dependency: openpyxl", parent=self)
			return

		SelectedFile = tkinter.filedialog.askopenfilename(
			title="Select XLSX inventory file",
			filetypes=(("Excel workbook", "*.xlsx"), ("All files", "*.*")),
			parent=self,
		)
		if not SelectedFile:
			return
		self._ProcessXlsxQueue([Path(SelectedFile).expanduser().resolve()])

	def _SelectAndProcessXlsxFolder(self) -> None:
		if openpyxl is None:
			tkinter.messagebox.showerror(WINDOW_TITLE, "Missing dependency: openpyxl", parent=self)
			return

		SelectedFolder = tkinter.filedialog.askdirectory(title="Select folder with XLSX files", parent=self)
		if not SelectedFolder:
			return

		Files = dgm_inventory.FindXlsxFiles(Path(SelectedFolder).expanduser().resolve())
		if not Files:
			tkinter.messagebox.showinfo(WINDOW_TITLE, "No .xlsx files found in the selected folder.", parent=self)
			return
		self._ProcessXlsxQueue(Files)

	def _ProcessXlsxQueue(self, Files: List[Path], Index: int = 0) -> None:
		if Index >= len(Files):
			tkinter.messagebox.showinfo(WINDOW_TITLE, "All selected XLSX files were processed.", parent=self)
			return

		try:
			Result = self._ProcessXlsxFile(Files[Index])
		except Exception as Error:
			tkinter.messagebox.showerror(WINDOW_TITLE, f"Cannot process '{Files[Index]}': {Error}", parent=self)
			return

		XlsxReviewWindow(self, Result, Files, Index)

	def _ProcessXlsxFile(self, FilePath: Path) -> GuiProcessResult:
		if openpyxl is None:
			raise RuntimeError("Missing dependency: openpyxl")

		Workbook = openpyxl.load_workbook(FilePath, data_only=False)
		Sheet = Workbook.active

		ProcessedRows: List[int] = []
		IgnoredRows = 0
		LastProcessedRow = 0
		ConsecutiveIgnoredRows = 0
		MissingByKey: Dict[Tuple[str, int, str], GuiMissingElement] = {}
		Conflicts: List[GuiConflictRow] = []
		Row = 1
		MaxRow = Sheet.max_row or 1

		while Row <= MaxRow:
			RawName = Sheet[f"{self.Database.Columns.Name}{Row}"].value
			if not dgm_inventory.CellHasUsableText(RawName):
				IgnoredRows += 1
				ConsecutiveIgnoredRows += 1
			else:
				Name = " ".join(str(RawName).strip().split())
				if self.Database.IsIgnoredText(Name):
					IgnoredRows += 1
					ConsecutiveIgnoredRows += 1
				else:
					SearchResult = self.Database.FindElement(Name)
					Entry = SearchResult.Record
					if Entry is None:
						MissingByKey[(Sheet.title, Row, Name)] = GuiMissingElement(FilePath, Sheet.title, Row, Name)
						ConsecutiveIgnoredRows = 0
					elif self._RowHasConflictingDgmValues(Sheet, Row, Entry):
						Conflicts.append(self._BuildConflict(FilePath, Sheet, Row, Name, Entry))
						ConsecutiveIgnoredRows = 0
					else:
						dgm_inventory.WriteEntryToRow(Sheet, Row, self.Database.Columns, Entry)
						ProcessedRows.append(Row)
						LastProcessedRow = Row
						ConsecutiveIgnoredRows = 0

			if ConsecutiveIgnoredRows >= dgm_inventory.STOP_AFTER_CONSECUTIVE_IGNORED_ROWS:
				break
			Row += 1

		TotalRow = LastProcessedRow + 1 if LastProcessedRow > 0 else 1
		dgm_inventory.WriteWorkbookTotals(Sheet, TotalRow, self.Database.Columns, ProcessedRows)
		Workbook.save(FilePath)
		return GuiProcessResult(FilePath, len(ProcessedRows), IgnoredRows, list(MissingByKey.values()), Conflicts)

	def _RowHasConflictingDgmValues(self, Sheet: object, Row: int, Entry: dgm_database.ElementRecord) -> bool:
		for MetalKey, _ in dgm_database.METALS:
			Value = self._ReadSheetDgmValue(Sheet, self.Database.Columns.PerElement[MetalKey], Row)
			if Value is not None and Value != 0 and Value != Entry.GetMetalValue(MetalKey):
				return True
		return False

	def _ReadSheetDgmValue(self, Sheet: object, Column: str, Row: int) -> Optional[decimal.Decimal]:
		Value = Sheet[f"{Column}{Row}"].value  # type: ignore[index]
		if Value is None or Value == "":
			return None
		try:
			return dgm_database.ReadDecimal(str(Value))
		except decimal.InvalidOperation:
			return decimal.Decimal("-1")

	def _BuildConflict(self, FilePath: Path, Sheet: object, Row: int, Name: str, Entry: dgm_database.ElementRecord) -> GuiConflictRow:
		SheetValues = dgm_database.DgmValues(decimal.Decimal("0"), decimal.Decimal("0"), decimal.Decimal("0"), decimal.Decimal("0"))
		Details: List[str] = []
		for MetalKey, MetalName in dgm_database.METALS:
			Value = self._ReadSheetDgmValue(Sheet, self.Database.Columns.PerElement[MetalKey], Row) or decimal.Decimal("0")
			SheetValues.SetMetalValue(MetalKey, Value)
			DbValue = Entry.GetMetalValue(MetalKey)
			if Value != 0 and Value != DbValue:
				Details.append(f"{MetalName}: sheet {dgm_database.DecimalToText(Value)} g, database {dgm_database.DecimalToText(DbValue)} g")
		DatabaseValues = dgm_database.DgmValues(Entry.Values.GoldG, Entry.Values.SilverG, Entry.Values.PlatinumG, Entry.Values.MpgG)
		return GuiConflictRow(FilePath, str(Sheet.title), Row, Name, Entry, SheetValues, DatabaseValues, "; ".join(Details))

	def _PopulateIgnoredList(self) -> None:
		self.IgnoredList.delete(0, tk.END)
		IgnoredValues = sorted(
			(Node.get("value", "") for Node in self.Database.IgnoredNode.findall("text")),
			key=dgm_database.NormalizeText,
		)
		for Value in IgnoredValues:
			self.IgnoredList.insert(tk.END, Value)
		self.IgnoredCountLabel.configure(text=f"{len(IgnoredValues)} ignored items")


class CatalogNodeEditDialog(tk.Toplevel):
	def __init__(self, Parent: tk.Tk, Node: XmlTree.Element) -> None:
		super().__init__(Parent)
		self.Result: Optional[Dict[str, str]] = None
		self.title("Modify database node")
		self.transient(Parent)
		self.grab_set()
		self.resizable(False, False)
		self.columnconfigure(1, weight=1)

		self._Entries: Dict[str, tk.Entry] = {}
		Row = 0
		for Key, Label, Value in self._BuildFields(Node):
			ttk.Label(self, text=Label).grid(row=Row, column=0, sticky="w", padx=(10, 6), pady=4)
			Entry = ttk.Entry(self, width=42)
			Entry.insert(0, Value)
			Entry.grid(row=Row, column=1, sticky="ew", padx=(0, 10), pady=4)
			self._Entries[Key] = Entry
			Row += 1

		ButtonFrame = ttk.Frame(self)
		ButtonFrame.grid(row=Row, column=0, columnspan=2, sticky="e", padx=10, pady=(8, 10))
		ttk.Button(ButtonFrame, text="Cancel", command=self._Cancel).grid(row=0, column=0, padx=(0, 6))
		ttk.Button(ButtonFrame, text="Save", command=self._Save).grid(row=0, column=1)

		self.bind("<Escape>", lambda _Event: self._Cancel())
		self.bind("<Return>", lambda _Event: self._Save())
		self.protocol("WM_DELETE_WINDOW", self._Cancel)
		self._Entries["text"].focus_set()
		self.wait_window(self)

	def _BuildFields(self, Node: XmlTree.Element) -> list[tuple[str, str, str]]:
		DgmNode = Node.find("dgm")
		SourceNode = DgmNode if DgmNode is not None else Node
		Fields = [
			("text", "Display text", Node.get("text", Node.get("name", ""))),
			("name", "Name", Node.get("name", "")),
		]
		if Node.tag == "regex":
			Fields.append(("pattern", "Regex pattern", Node.get("pattern", "")))
		for MetalKey, MetalName in dgm_database.METALS:
			Fields.append((MetalKey, f"{MetalName} g", SourceNode.get(f"{MetalKey}_g", "0")))
		return Fields

	def _Save(self) -> None:
		try:
			for MetalKey, _ in dgm_database.METALS:
				dgm_database.ReadDecimal(self._Entries[MetalKey].get())
		except decimal.InvalidOperation as Error:
			tkinter.messagebox.showerror(WINDOW_TITLE, f"Invalid decimal value: {Error}", parent=self)
			return

		self.Result = {Key: Entry.get() for Key, Entry in self._Entries.items()}
		self.destroy()

	def _Cancel(self) -> None:
		self.Result = None
		self.destroy()


class MoveCatalogNodeDialog(tk.Toplevel):
	def __init__(self, Parent: tk.Tk, CurrentParentPath: List[str]) -> None:
		super().__init__(Parent)
		self.Result: Optional[List[str]] = None

		self.title("Move catalog node")
		self.transient(Parent)
		self.grab_set()
		self.resizable(False, False)

		self.columnconfigure(0, weight=1)
		self.rowconfigure(1, weight=1)

		ttk.Label(self, text="New parent path").grid(row=0, column=0, sticky="w", padx=10, pady=(10, 4))

		self.PathList = tk.Listbox(self, height=8, width=48, exportselection=False)
		self.PathList.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 6))

		for Part in CurrentParentPath:
			self.PathList.insert(tk.END, Part)

		EditFrame = ttk.Frame(self)
		EditFrame.grid(row=2, column=0, sticky="ew", padx=10, pady=(0, 6))
		EditFrame.columnconfigure(0, weight=1)

		self.PartEntry = ttk.Entry(EditFrame)
		self.PartEntry.grid(row=0, column=0, sticky="ew", padx=(0, 6))

		ttk.Button(EditFrame, text="Add", command=self._AddPart).grid(row=0, column=1, padx=(0, 6))
		ttk.Button(EditFrame, text="Replace", command=self._ReplacePart).grid(row=0, column=2)

		ButtonFrame = ttk.Frame(self)
		ButtonFrame.grid(row=3, column=0, sticky="ew", padx=10, pady=(0, 10))

		ttk.Button(ButtonFrame, text="Remove", command=self._RemovePart).grid(row=0, column=0, padx=(0, 6))
		ttk.Button(ButtonFrame, text="Up", command=self._MovePartUp).grid(row=0, column=1, padx=(0, 6))
		ttk.Button(ButtonFrame, text="Down", command=self._MovePartDown).grid(row=0, column=2, padx=(0, 6))

		ttk.Button(ButtonFrame, text="Cancel", command=self._Cancel).grid(row=0, column=3, sticky="e", padx=(20, 6))
		ttk.Button(ButtonFrame, text="Move", command=self._Move).grid(row=0, column=4, sticky="e")

		ttk.Label(
			self,
			text="Each row is one parent node. Empty list means catalog root."
		).grid(row=4, column=0, sticky="w", padx=10, pady=(0, 10))

		self.PathList.bind("<<ListboxSelect>>", self._OnSelect)

		self.bind("<Escape>", lambda _Event: self._Cancel())
		self.bind("<Return>", lambda _Event: self._Move())
		self.protocol("WM_DELETE_WINDOW", self._Cancel)

		self.PartEntry.focus_set()
		self.wait_window(self)

	def _GetSelectedIndex(self) -> Optional[int]:
		Selection = self.PathList.curselection()
		if not Selection:
			return None
		return int(Selection[0])

	def _OnSelect(self, _Event: tk.Event) -> None:
		Index = self._GetSelectedIndex()
		if Index is None:
			return

		self.PartEntry.delete(0, tk.END)
		self.PartEntry.insert(0, self.PathList.get(Index))

	def _AddPart(self) -> None:
		Part = self.PartEntry.get()

		# ВАЖЛИВО: не робимо .strip(), бо trailing space може бути значущим.
		if Part == "":
			return

		self.PathList.insert(tk.END, Part)
		self.PartEntry.delete(0, tk.END)

	def _ReplacePart(self) -> None:
		Index = self._GetSelectedIndex()
		if Index is None:
			return

		Part = self.PartEntry.get()

		# ВАЖЛИВО: не робимо .strip().
		if Part == "":
			return

		self.PathList.delete(Index)
		self.PathList.insert(Index, Part)
		self.PathList.selection_set(Index)

	def _RemovePart(self) -> None:
		Index = self._GetSelectedIndex()
		if Index is None:
			return

		self.PathList.delete(Index)

	def _MovePartUp(self) -> None:
		Index = self._GetSelectedIndex()
		if Index is None or Index <= 0:
			return

		Part = self.PathList.get(Index)
		self.PathList.delete(Index)
		self.PathList.insert(Index - 1, Part)
		self.PathList.selection_set(Index - 1)

	def _MovePartDown(self) -> None:
		Index = self._GetSelectedIndex()
		if Index is None or Index >= self.PathList.size() - 1:
			return

		Part = self.PathList.get(Index)
		self.PathList.delete(Index)
		self.PathList.insert(Index + 1, Part)
		self.PathList.selection_set(Index + 1)

	def _Move(self) -> None:
		self.Result = []

		for Index in range(self.PathList.size()):
			self.Result.append(self.PathList.get(Index))

		self.destroy()

	def _Cancel(self) -> None:
		self.Result = None
		self.destroy()


class XlsxReviewWindow(tk.Toplevel):
	def __init__(self, Parent: DgmDatabaseViewer, Result: GuiProcessResult, Files: List[Path], Index: int) -> None:
		super().__init__(Parent)
		self.ParentViewer = Parent
		self.Result = Result
		self.Files = Files
		self.Index = Index
		self.MissingItems = list(Result.Missing)
		self.ConflictItems = list(Result.Conflicts)

		self.title(f"XLSX review - {Result.FilePath.name}")
		self.geometry("1100x620")
		self.minsize(800, 420)
		self.columnconfigure(0, weight=1)
		self.rowconfigure(1, weight=1)

		ttk.Label(
			self,
			text=(
				f"Processed {Result.ProcessedRows} rows, ignored {Result.IgnoredRows} rows. "
				f"Missing: {len(Result.Missing)}. Conflicts: {len(Result.Conflicts)}."
			),
			style="Heading.TLabel",
		).grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 4))

		Content = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
		Content.grid(row=1, column=0, sticky="nsew", padx=10, pady=6)
		MissingPane = ttk.Frame(Content)
		ConflictPane = ttk.Frame(Content)
		Content.add(MissingPane, weight=1)
		Content.add(ConflictPane, weight=1)
		self._BuildMissingPane(MissingPane)
		self._BuildConflictPane(ConflictPane)

		ButtonFrame = ttk.Frame(self)
		ButtonFrame.grid(row=2, column=0, sticky="e", padx=10, pady=(4, 10))
		ttk.Button(ButtonFrame, text="Process current file again", command=self._ProcessAgain).grid(row=0, column=0, padx=(0, 6))
		if Index + 1 < len(Files):
			ttk.Button(ButtonFrame, text="Process next file", command=self._ProcessNext).grid(row=0, column=1, padx=(0, 6))
		ttk.Button(ButtonFrame, text="Close", command=self.destroy).grid(row=0, column=2)

	def _BuildMissingPane(self, Parent: ttk.Frame) -> None:
		Parent.columnconfigure(0, weight=1)
		Parent.rowconfigure(1, weight=1)
		ttk.Label(Parent, text="Missing database elements", style="Heading.TLabel").grid(row=0, column=0, sticky="w")
		self.MissingTree = ttk.Treeview(Parent, columns=("row", "name"), show="headings", selectmode="browse")
		self.MissingTree.heading("row", text="Row")
		self.MissingTree.heading("name", text="Element text")
		self.MissingTree.column("row", width=70, stretch=False, anchor="e")
		self.MissingTree.column("name", width=360, stretch=True)
		self.MissingTree.grid(row=1, column=0, sticky="nsew", pady=4)
		for Index, Item in enumerate(self.MissingItems):
			self.MissingTree.insert("", "end", iid=str(Index), values=(Item.Row, Item.Name))
		Buttons = ttk.Frame(Parent)
		Buttons.grid(row=2, column=0, sticky="ew")
		ttk.Button(Buttons, text="Add to database...", command=self._AddMissingToDatabase).grid(row=0, column=0, padx=(0, 6))
		ttk.Button(Buttons, text="Add to ignore list", command=self._IgnoreMissing).grid(row=0, column=1, padx=(0, 6))
		ttk.Button(Buttons, text="Rename in XLSX...", command=self._RenameMissing).grid(row=0, column=2)

	def _BuildConflictPane(self, Parent: ttk.Frame) -> None:
		Parent.columnconfigure(0, weight=1)
		Parent.rowconfigure(1, weight=1)
		ttk.Label(Parent, text="Rows conflicting with database", style="Heading.TLabel").grid(row=0, column=0, sticky="w")
		self.ConflictTree = ttk.Treeview(Parent, columns=("row", "name", "details"), show="headings", selectmode="browse")
		for Column, Label, Width in (("row", "Row", 70), ("name", "Element text", 220), ("details", "Conflict", 380)):
			self.ConflictTree.heading(Column, text=Label)
			self.ConflictTree.column(Column, width=Width, stretch=(Column == "details"))
		self.ConflictTree.grid(row=1, column=0, sticky="nsew", pady=4)
		for Index, Item in enumerate(self.ConflictItems):
			self.ConflictTree.insert("", "end", iid=str(Index), values=(Item.Row, Item.Name, Item.Details))
		Buttons = ttk.Frame(Parent)
		Buttons.grid(row=2, column=0, sticky="ew")
		ttk.Button(Buttons, text="Clean XLSX row", command=self._CleanConflictRow).grid(row=0, column=0, padx=(0, 6))
		ttk.Button(Buttons, text="Update database from row", command=self._UpdateDatabaseFromConflict).grid(row=0, column=1)

	def _SelectedMissing(self) -> Optional[GuiMissingElement]:
		Selection = self.MissingTree.selection()
		return self.MissingItems[int(Selection[0])] if Selection else None

	def _SelectedConflict(self) -> Optional[GuiConflictRow]:
		Selection = self.ConflictTree.selection()
		return self.ConflictItems[int(Selection[0])] if Selection else None

	def _AddMissingToDatabase(self) -> None:
		Item = self._SelectedMissing()
		if Item is None:
			return
		Dialog = AddElementDialog(self, self.ParentViewer.Database, Item.Name)
		if Dialog.Result is None:
			return
		try:
			Mode, Values, PathParts = Dialog.Result
			if Mode == "existing":
				self.ParentViewer.Database.AddDgmToExistingPath(Item.Name, Values, PathParts)
			else:
				self.ParentViewer.Database.AddElement(Item.Name, Values, PathParts)
			self.ParentViewer.Database.Save()
			self.ParentViewer._PopulateDatabaseViews()
			self.destroy()
			self.ParentViewer._ProcessXlsxQueue(self.Files, self.Index)
		except Exception as Error:
			tkinter.messagebox.showerror(WINDOW_TITLE, str(Error), parent=self)

	def _IgnoreMissing(self) -> None:
		Item = self._SelectedMissing()
		if Item is None:
			return
		self.ParentViewer.Database.AddIgnoredText(Item.Name)
		self.ParentViewer.Database.Save()
		self.ParentViewer._PopulateIgnoredList()
		self.destroy()
		self.ParentViewer._ProcessXlsxQueue(self.Files, self.Index)

	def _RenameMissing(self) -> None:
		Item = self._SelectedMissing()
		if Item is None:
			return
		Dialog = RenameElementDialog(self, Item.Name)
		if Dialog.Result is None:
			return
		Workbook = openpyxl.load_workbook(Item.FilePath, data_only=False)  # type: ignore[union-attr]
		Workbook[Item.SheetName][f"{self.ParentViewer.Database.Columns.Name}{Item.Row}"].value = Dialog.Result
		Workbook.save(Item.FilePath)
		self.destroy()
		self.ParentViewer._ProcessXlsxQueue(self.Files, self.Index)

	def _CleanConflictRow(self) -> None:
		Item = self._SelectedConflict()
		if Item is None:
			return
		Workbook = openpyxl.load_workbook(Item.FilePath, data_only=False)  # type: ignore[union-attr]
		Sheet = Workbook[Item.SheetName]
		for MetalKey, _ in dgm_database.METALS:
			Sheet[f"{self.ParentViewer.Database.Columns.PerElement[MetalKey]}{Item.Row}"].value = None
			Sheet[f"{self.ParentViewer.Database.Columns.Total[MetalKey]}{Item.Row}"].value = None
		Workbook.save(Item.FilePath)
		self.destroy()
		self.ParentViewer._ProcessXlsxQueue(self.Files, self.Index)

	def _UpdateDatabaseFromConflict(self) -> None:
		Item = self._SelectedConflict()
		if Item is None:
			return
		Item.Record.SetValues(Item.SheetValues)
		self.ParentViewer.Database.Save()
		self.ParentViewer._PopulateDatabaseViews()
		self.destroy()
		self.ParentViewer._ProcessXlsxQueue(self.Files, self.Index)

	def _ProcessAgain(self) -> None:
		self.destroy()
		self.ParentViewer._ProcessXlsxQueue(self.Files, self.Index)

	def _ProcessNext(self) -> None:
		self.destroy()
		self.ParentViewer._ProcessXlsxQueue(self.Files, self.Index + 1)


class RenameElementDialog(tk.Toplevel):
	def __init__(self, Parent: tk.Toplevel, CurrentName: str) -> None:
		super().__init__(Parent)
		self.Result: Optional[str] = None
		self.title("Rename XLSX element")
		self.transient(Parent)
		self.grab_set()
		self.resizable(False, False)
		ttk.Label(self, text="New element text").grid(row=0, column=0, sticky="w", padx=10, pady=(10, 4))
		self.Entry = ttk.Entry(self, width=60)
		self.Entry.insert(0, CurrentName)
		self.Entry.grid(row=1, column=0, sticky="ew", padx=10)
		Buttons = ttk.Frame(self)
		Buttons.grid(row=2, column=0, sticky="e", padx=10, pady=10)
		ttk.Button(Buttons, text="Cancel", command=self._Cancel).grid(row=0, column=0, padx=(0, 6))
		ttk.Button(Buttons, text="Save", command=self._Save).grid(row=0, column=1)
		self.wait_window(self)

	def _Save(self) -> None:
		Value = self.Entry.get().strip()
		if Value:
			self.Result = Value
			self.destroy()

	def _Cancel(self) -> None:
		self.Result = None
		self.destroy()


class AddElementDialog(tk.Toplevel):
	def __init__(self, Parent: tk.Toplevel, Db: dgm_database.DgmDatabase, Name: str) -> None:
		super().__init__(Parent)
		self.Result: Optional[Tuple[str, dgm_database.DgmValues, List[str]]] = None
		self.Db = Db
		self.Name = Name
		self.Candidates = [Candidate for Candidate in Db.FindExistingRegularPathCandidates(Name) if not Candidate.HasDgm]
		self.title("Add missing element")
		self.transient(Parent)
		self.grab_set()
		self.geometry("560x520")
		self.columnconfigure(0, weight=1)
		self.rowconfigure(4, weight=1)
		ttk.Label(self, text=f"Element: {Name}", style="Heading.TLabel").grid(row=0, column=0, sticky="w", padx=10, pady=(10, 4))
		self.UseCandidate = tk.BooleanVar(value=bool(self.Candidates))
		ttk.Checkbutton(self, text="Add DGM to selected partial/existing path candidate", variable=self.UseCandidate).grid(row=1, column=0, sticky="w", padx=10)
		self.CandidateList = tk.Listbox(self, height=4, exportselection=False)
		self.CandidateList.grid(row=2, column=0, sticky="ew", padx=10, pady=4)
		for Candidate in self.Candidates:
			self.CandidateList.insert(tk.END, Candidate.DisplayPath)
		if self.Candidates:
			self.CandidateList.selection_set(0)
		ttk.Label(self, text="Structured node chain (one node per line)").grid(row=3, column=0, sticky="w", padx=10, pady=(8, 4))
		self.PathList = tk.Listbox(self, height=8, exportselection=False)
		self.PathList.grid(row=4, column=0, sticky="nsew", padx=10)
		for Part in self._DefaultSplit(Name):
			self.PathList.insert(tk.END, Part)
		Edit = ttk.Frame(self)
		Edit.grid(row=5, column=0, sticky="ew", padx=10, pady=4)
		Edit.columnconfigure(0, weight=1)
		self.PartEntry = ttk.Entry(Edit)
		self.PartEntry.grid(row=0, column=0, sticky="ew", padx=(0, 6))
		ttk.Button(Edit, text="Split by /", command=self._SplitEntry).grid(row=0, column=1, padx=(0, 6))
		ttk.Button(Edit, text="Add", command=self._AddPart).grid(row=0, column=2, padx=(0, 6))
		ttk.Button(Edit, text="Replace", command=self._ReplacePart).grid(row=0, column=3)
		Controls = ttk.Frame(self)
		Controls.grid(row=6, column=0, sticky="w", padx=10)
		ttk.Button(Controls, text="Remove", command=self._RemovePart).grid(row=0, column=0, padx=(0, 6))
		ttk.Button(Controls, text="Use candidate as parent", command=self._UseCandidateAsParent).grid(row=0, column=1)
		Values = ttk.LabelFrame(self, text="DGM values, g")
		Values.grid(row=7, column=0, sticky="ew", padx=10, pady=8)
		self.ValueEntries: Dict[str, ttk.Entry] = {}
		for Index, (MetalKey, MetalName) in enumerate(dgm_database.METALS):
			ttk.Label(Values, text=MetalName).grid(row=0, column=Index, sticky="w")
			Entry = ttk.Entry(Values, width=12)
			Entry.insert(0, "0")
			Entry.grid(row=1, column=Index, padx=(0, 6))
			self.ValueEntries[MetalKey] = Entry
		Buttons = ttk.Frame(self)
		Buttons.grid(row=8, column=0, sticky="e", padx=10, pady=(0, 10))
		ttk.Button(Buttons, text="Cancel", command=self._Cancel).grid(row=0, column=0, padx=(0, 6))
		ttk.Button(Buttons, text="Add", command=self._Save).grid(row=0, column=1)
		self.wait_window(self)

	def _DefaultSplit(self, Name: str) -> List[str]:
		return [Part for Part in Name.split("/") if Part] or [Name]

	def _GetCandidate(self) -> Optional[dgm_database.ExistingPathInfo]:
		Selection = self.CandidateList.curselection()
		return self.Candidates[int(Selection[0])] if Selection else None

	def _UseCandidateAsParent(self) -> None:
		Candidate = self._GetCandidate()
		if Candidate is None:
			return
		self.PathList.delete(0, tk.END)
		for Part in Candidate.PathParts + [self.Name]:
			self.PathList.insert(tk.END, Part)
		self.UseCandidate.set(False)

	def _SplitEntry(self) -> None:
		Parts = [Part for Part in self.PartEntry.get().split("/") if Part]
		if not Parts:
			Parts = self._DefaultSplit(self.Name)
		for Part in Parts:
			self.PathList.insert(tk.END, Part)
		self.PartEntry.delete(0, tk.END)

	def _AddPart(self) -> None:
		if self.PartEntry.get():
			self.PathList.insert(tk.END, self.PartEntry.get())
			self.PartEntry.delete(0, tk.END)

	def _ReplacePart(self) -> None:
		Selection = self.PathList.curselection()
		if Selection and self.PartEntry.get():
			Index = int(Selection[0])
			self.PathList.delete(Index)
			self.PathList.insert(Index, self.PartEntry.get())

	def _RemovePart(self) -> None:
		Selection = self.PathList.curselection()
		if Selection:
			self.PathList.delete(int(Selection[0]))

	def _Save(self) -> None:
		try:
			Values = dgm_database.DgmValues(
				GoldG=dgm_database.ReadDecimal(self.ValueEntries["gold"].get()),
				SilverG=dgm_database.ReadDecimal(self.ValueEntries["silver"].get()),
				PlatinumG=dgm_database.ReadDecimal(self.ValueEntries["platinum"].get()),
				MpgG=dgm_database.ReadDecimal(self.ValueEntries["mpg"].get()),
			)
		except decimal.InvalidOperation as Error:
			tkinter.messagebox.showerror(WINDOW_TITLE, f"Invalid DGM value: {Error}", parent=self)
			return
		Candidate = self._GetCandidate() if self.UseCandidate.get() else None
		if Candidate is not None:
			self.Result = ("existing", Values, Candidate.PathParts)
		else:
			PathParts = [self.PathList.get(Index) for Index in range(self.PathList.size())]
			if not PathParts:
				tkinter.messagebox.showerror(WINDOW_TITLE, "Enter at least one node.", parent=self)
				return
			self.Result = ("new", Values, PathParts)
		self.destroy()

	def _Cancel(self) -> None:
		self.Result = None
		self.destroy()


def SelectDatabaseWithDialog() -> Optional[Path]:
	Root = tk.Tk()
	Root.withdraw()
	Root.update()
	try:
		SelectedFile = tkinter.filedialog.askopenfilename(
			title="Select DGM XML database",
			filetypes=(("XML database", "*.xml"), ("All files", "*.*")),
		)
	finally:
		Root.destroy()

	if not SelectedFile:
		return None
	return Path(SelectedFile).expanduser().resolve()


def ResolveDatabaseArgument(Argument: Optional[str]) -> Optional[Path]:
	if Argument is not None:
		return Path(Argument).expanduser().resolve()
	return SelectDatabaseWithDialog()


def Main() -> int:
	if len(sys.argv) not in (1, 2):
		print("Usage: python dgm_gui.py [database.xml]", file=sys.stderr)
		return 2

	DatabasePath = ResolveDatabaseArgument(sys.argv[1] if len(sys.argv) == 2 else None)
	if DatabasePath is None:
		print("No database selected.", file=sys.stderr)
		return 2
	if not DatabasePath.exists() or not DatabasePath.is_file():
		print(f"Database file does not exist: {DatabasePath}", file=sys.stderr)
		return 2

	try:
		Application = DgmDatabaseViewer(DatabasePath)
	except Exception as Error:
		try:
			tkinter.messagebox.showerror(WINDOW_TITLE, str(Error))
		except Exception:
			pass
		print(str(Error), file=sys.stderr)
		return 1

	Application.mainloop()
	return 0


if __name__ == "__main__":
	sys.exit(Main())
