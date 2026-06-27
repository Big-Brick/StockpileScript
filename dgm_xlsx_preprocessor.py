from __future__ import annotations

import re
import xml.etree.ElementTree as XmlTree
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Pattern, Sequence, Tuple

import dgm_database
import dgm_xlsx_common

try:
	import openpyxl
except ImportError:
	openpyxl = None  # type: ignore[assignment]


RULES_VERSION = "1"
DEFAULT_RULES_FILENAME = "preprocess_rules.xml"
DATABASE_PREFIX_REGEX_ATTR = "preprocess_prefix_regex"


@dataclass
class PreprocessRegexRule:
	Id: str
	Pattern: str
	Replacement: str
	Description: str = ""
	Enabled: bool = True
	ElementTypes: List[str] = field(default_factory=list)


@dataclass
class PreprocessElementType:
	Id: str
	Canonical: str
	Aliases: List[str] = field(default_factory=list)
	DatabaseCheck: bool = True
	XmlNode: Optional[XmlTree.Element] = None


@dataclass
class RuntimePrefixRule:
	ElementType: PreprocessElementType
	Pattern: Pattern[str]


@dataclass
class PreprocessRules:
	Path: Path
	Tree: XmlTree.ElementTree
	Root: XmlTree.Element
	IgnoreCase: bool = True
	CollapseWhitespace: bool = True
	TrimWhitespace: bool = True
	IgnoredTextRules: List[PreprocessRegexRule] = field(default_factory=list)
	StageOneRules: List[PreprocessRegexRule] = field(default_factory=list)
	StageThreeRules: List[PreprocessRegexRule] = field(default_factory=list)
	ElementTypes: List[PreprocessElementType] = field(default_factory=list)


@dataclass
class PreprocessChange:
	Row: int
	OriginalText: str
	NewText: str
	StageNotes: List[str]
	DatabaseVerified: bool = False
	Ambiguous: bool = False
	StageId: str = ""
	ElementType: str = ""
	SafetyLevel: str = ""
	FilePath: Path = Path()
	SheetName: str = ""
	RulesPath: Path = Path()
	StageName: str = "All stages"


@dataclass
class PreprocessStage:
	Id: str
	Name: str


PREPROCESS_STAGES: Tuple[PreprocessStage, ...] = (
	PreprocessStage("language", "Language normalization"),
	PreprocessStage("prefix", "Add element prefixes"),
	PreprocessStage("technical", "Technical normalization"),
)

PREFIX_SAFETY_ORDER = {"safe": 0, "partial": 1, "ambiguous": 2, "unidentified": 3}
PREFIX_SAFETY_LABELS = {
	"safe": "Safe exact database match",
	"partial": "Partially safe partial database match",
	"ambiguous": "Ambiguous element type",
	"unidentified": "Unidentified by rules",
}

# Stage 1 character-level normalization. These are Latin/Russian glyphs that
# are routinely typed into Ukrainian/Cyrillic element designations by OCR or
# mixed keyboard layout and should be normalized before regex rules run.
UKRAINIAN_LOOKALIKE_TRANSLATION = str.maketrans({
	"A": "А", "B": "В", "C": "С", "E": "Е", "H": "Н", "I": "І", "K": "К", "M": "М", "O": "О", "P": "Р", "T": "Т", "X": "Х", "Y": "У",
	"a": "а", "c": "с", "e": "е", "i": "і", "k": "к", "m": "м", "o": "о", "p": "р", "t": "т", "x": "х", "y": "у",
	"Ё": "Е", "ё": "е", "Э": "Е", "э": "е",
})


class XlsxPreprocessor:
	def __init__(self, Database: dgm_database.DgmDatabase, RulesPath: Path) -> None:
		self.Database = Database
		self.Rules = LoadPreprocessRules(RulesPath)
		self.RegexFlags = re.IGNORECASE if self.Rules.IgnoreCase else 0
		self.RuntimePrefixRules = self._BuildRuntimePrefixRules()

	def PreprocessWorkbook(self, FilePath: Path, StageId: str = "all") -> List[PreprocessChange]:
		if openpyxl is None:
			raise RuntimeError("Missing dependency: openpyxl")

		Workbook = openpyxl.load_workbook(FilePath, data_only=False)
		Sheet = Workbook.active
		Changes: List[PreprocessChange] = []
		ConsecutiveIgnoredRows = 0
		Row = 1
		MaxRow = Sheet.max_row or 1

		while Row <= MaxRow:
			RawName = Sheet[f"{self.Database.Columns.Name}{Row}"].value
			if not dgm_xlsx_common.CellHasUsableText(RawName):
				ConsecutiveIgnoredRows += 1
			else:
				Original = str(RawName)
				if self.IsIgnoredText(Original):
					ConsecutiveIgnoredRows += 1
				else:
					Change = self.PreprocessText(Row, Original, StageId)
					self._AttachChangeMetadata(Change, FilePath, Sheet.title, StageId)
					if self._ShouldOfferChange(Change, StageId):
						Changes.append(Change)
					ConsecutiveIgnoredRows = 0

			if ConsecutiveIgnoredRows >= dgm_xlsx_common.STOP_AFTER_CONSECUTIVE_IGNORED_ROWS:
				break
			Row += 1

		return Changes

	def _AttachChangeMetadata(self, Change: PreprocessChange, FilePath: Path, SheetName: str, StageId: str) -> PreprocessChange:
		Change.FilePath = FilePath
		Change.SheetName = SheetName
		Change.RulesPath = self.Rules.Path
		Change.StageId = StageId
		Change.StageName = self.GetStageName(StageId)
		return Change

	def IsIgnoredText(self, Text: str) -> bool:
		Current = self._CleanWhitespace(Text)
		for Rule in self.Rules.IgnoredTextRules:
			if not Rule.Enabled or not Rule.Pattern:
				continue
			try:
				if re.search(Rule.Pattern, Current, flags=self.RegexFlags):
					return True
			except re.error:
				continue
		return False

	def ApplyChanges(self, Changes: Sequence[PreprocessChange]) -> None:
		if openpyxl is None:
			raise RuntimeError("Missing dependency: openpyxl")
		if not Changes:
			return

		ChangesBySheet: Dict[Tuple[Path, str], List[PreprocessChange]] = {}
		for Change in Changes:
			ChangesBySheet.setdefault((Change.FilePath, Change.SheetName), []).append(Change)

		for (FilePath, SheetName), SheetChanges in ChangesBySheet.items():
			Workbook = openpyxl.load_workbook(FilePath, data_only=False)
			Sheet = Workbook[SheetName]
			for Change in SheetChanges:
				Sheet[f"{self.Database.Columns.Name}{Change.Row}"].value = Change.NewText
			Workbook.save(FilePath)

	def PreprocessText(self, Row: int, Text: str, StageId: str = "all") -> PreprocessChange:
		if StageId == "language":
			return self._PreprocessLanguage(Row, Text)
		if StageId == "prefix":
			return self._PreprocessPrefix(Row, Text)
		if StageId == "technical":
			return self._PreprocessTechnical(Row, Text)

		Notes: List[str] = []
		Original = Text
		Current = self._CleanWhitespace(Text)
		if Current != Text:
			Notes.append("Cleaned whitespace")
		if self.IsIgnoredText(Current):
			Notes.append("Ignored block/header row")
			return PreprocessChange(Row, Original, Current, Notes, False, False)

		Current = self._NormalizeUkrainianLookalikes(Current, Notes)
		Current = self._ApplyRules(Current, self.Rules.StageOneRules, "Stage 1", Notes)
		Current, TypeVerified, TypeAmbiguous = self._NormalizeType(Current, Notes)
		Current = self._ApplyTechnicalNormalization(Current, Notes)
		Current = self._ApplyExactDatabaseCasing(Current, Notes)

		DatabaseRecord = self.Database.FindElement(Current).Record
		DatabaseVerified = TypeVerified or (DatabaseRecord is not None and DatabaseRecord.HasDgm)
		if DatabaseVerified and not TypeVerified:
			Notes.append("Verified final text against database")
		if not Notes and Original == Current:
			Notes.append("No change")
		return PreprocessChange(Row, Original, Current, Notes, DatabaseVerified, TypeAmbiguous, "all")

	def _PreprocessLanguage(self, Row: int, Text: str) -> PreprocessChange:
		Notes: List[str] = []
		Original = Text
		Current = self._CleanWhitespace(Text)
		if Current != Text:
			Notes.append("Cleaned whitespace")
		Current = self._NormalizeUkrainianLookalikes(Current, Notes)
		Current = self._ApplyRules(Current, self.Rules.StageOneRules, "Stage 1", Notes)
		return PreprocessChange(Row, Original, Current, Notes, True, False, "language")

	def _PreprocessTechnical(self, Row: int, Text: str) -> PreprocessChange:
		Notes: List[str] = []
		Original = Text
		Current = self._CleanWhitespace(Text)
		if Current != Text:
			Notes.append("Cleaned whitespace")
		Current = self._ApplyTechnicalNormalization(Current, Notes)
		Current = self._ApplyExactDatabaseCasing(Current, Notes)
		DatabaseRecord = self.Database.FindElement(Current).Record
		Verified = DatabaseRecord is not None and DatabaseRecord.HasDgm
		return PreprocessChange(Row, Original, Current, Notes, Verified, False, "technical")

	def _ApplyExactDatabaseCasing(self, Text: str, Notes: List[str]) -> str:
		Result = self.Database.FindElement(Text)
		Record = Result.Record
		if Record is None or not Record.HasDgm or Result.MatchedByRegex:
			return Text
		DatabaseText = BuildExactDatabaseText(Record)
		if not DatabaseText:
			return Text
		if Text != DatabaseText and Text.casefold() == DatabaseText.casefold() and len(Text) == len(DatabaseText):
			Notes.append("Stage 3: matched exact database letter casing")
			return DatabaseText
		return Text

	def _ApplyTechnicalNormalization(self, Text: str, Notes: List[str]) -> str:
		Explicit = self._FindLeadingExplicitType(Text)
		if Explicit is None:
			return Text

		ElementType, Remainder = Explicit
		Current = self._ApplyRules(Remainder, self.Rules.StageThreeRules, "Stage 3", Notes, ElementType)
		CaseNormalized = NormalizeDesignationCase(Current)
		if CaseNormalized != Current:
			Notes.append("Stage 3: normalized designation letter casing")
			Current = CaseNormalized
		return self._CleanWhitespace(f"{ElementType.Canonical} {Current}")


	def _PreprocessPrefix(self, Row: int, Text: str) -> PreprocessChange:
		Original = Text
		Current = self._CleanWhitespace(Text)
		ExistingType = self._FindExplicitType(Current)
		ExistingResult = self.Database.FindElement(Current)
		ExistingRecord = ExistingResult.Record
		if (
			ExistingType is not None
			and ExistingRecord is not None
			and ExistingRecord.HasDgm
			and not ExistingResult.MatchedByRegex
			and self._RecordMatchesElementType(ExistingRecord, ExistingType[0])
		):
			return PreprocessChange(Row, Original, Current, ["Stage 2: existing prefixed text verified by exact database match"], True, False, "prefix", ExistingType[0].Canonical, "safe")
		elif ExistingType is not None:
			ElementType, Remainder = ExistingType
			Candidate = self._CleanWhitespace(f"{ElementType.Canonical} {Remainder}")
			Safety = self._ClassifyPrefixCandidate(Candidate, ElementType)
			if Safety in ("safe", "partial") and Candidate == Current:
				return PreprocessChange(Row, Original, Current, [f"Stage 2: existing prefix accepted ({PREFIX_SAFETY_LABELS[Safety]})"], Safety == "safe", False, "prefix", ElementType.Canonical, Safety)
			return PreprocessChange(Row, Original, Candidate, [f"Stage 2: normalized explicit type as {ElementType.Canonical}"], Safety == "safe", Safety == "ambiguous", "prefix", ElementType.Canonical, Safety)

		PrefixMatchesByPattern: List[Tuple[PreprocessElementType, str, str]] = []
		for PrefixRule in self.RuntimePrefixRules:
			if PrefixRule.Pattern.search(Current) is None:
				continue
			ElementType = PrefixRule.ElementType
			Candidate = self._MakeTypedCandidate(ElementType, Current)
			Safety = self._ClassifyPrefixCandidate(Candidate, ElementType)
			if Safety in ("safe", "partial"):
				PrefixMatchesByPattern.append((ElementType, Candidate, Safety))
		if len(PrefixMatchesByPattern) == 1:
			ElementType, Candidate, Safety = PrefixMatchesByPattern[0]
			return PreprocessChange(Row, Original, Candidate, [f"Stage 2: inferred {ElementType.Canonical} from database prefix pattern"], Safety == "safe", False, "prefix", ElementType.Canonical, Safety)
		if len(PrefixMatchesByPattern) > 1:
			Types = ", ".join(Match[0].Canonical for Match in PrefixMatchesByPattern)
			return PreprocessChange(Row, Original, Current, [f"Stage 2: ambiguous database prefix patterns: {Types}"], False, True, "prefix", "", "ambiguous")

		ExactMatches: List[Tuple[PreprocessElementType, str]] = []
		PartialMatches: List[Tuple[PreprocessElementType, str]] = []
		for ElementType in self.Rules.ElementTypes:
			Candidate = self._MakeTypedCandidate(ElementType, Current)
			Safety = self._ClassifyPrefixCandidate(Candidate, ElementType)
			if Safety == "safe":
				ExactMatches.append((ElementType, Candidate))
			elif Safety == "partial":
				PartialMatches.append((ElementType, Candidate))

		if len(ExactMatches) == 1:
			ElementType, Candidate = ExactMatches[0]
			return PreprocessChange(Row, Original, Candidate, [f"Stage 2: inferred {ElementType.Canonical} by exact database match"], True, False, "prefix", ElementType.Canonical, "safe")
		if len(ExactMatches) > 1:
			Types = ", ".join(Match[0].Canonical for Match in ExactMatches)
			return PreprocessChange(Row, Original, Current, [f"Stage 2: ambiguous exact database matches: {Types}"], False, True, "prefix", "", "ambiguous")
		if len(PartialMatches) == 1:
			ElementType, Candidate = PartialMatches[0]
			return PreprocessChange(Row, Original, Candidate, [f"Stage 2: inferred {ElementType.Canonical} by partial database match"], False, False, "prefix", ElementType.Canonical, "partial")
		if len(PartialMatches) > 1:
			Types = ", ".join(Match[0].Canonical for Match in PartialMatches)
			return PreprocessChange(Row, Original, Current, [f"Stage 2: ambiguous partial database matches: {Types}"], False, True, "prefix", "", "ambiguous")
		return PreprocessChange(Row, Original, Current, ["Stage 2: element type was not identified by rules"], False, True, "prefix", "", "unidentified")

	def _ShouldOfferChange(self, Change: PreprocessChange, StageId: str) -> bool:
		if StageId == "prefix":
			if Change.NewText == Change.OriginalText and Change.ElementType:
				return False
			return Change.SafetyLevel not in ("safe", "partial") or Change.NewText != Change.OriginalText
		return Change.NewText != Change.OriginalText

	def _ClassifyPrefixCandidate(self, Candidate: str, ElementType: PreprocessElementType) -> str:
		Result = self.Database.FindElement(Candidate)
		if (
			Result.Record is not None
			and Result.Record.HasDgm
			and not Result.MatchedByRegex
			and self._RecordMatchesElementType(Result.Record, ElementType)
		):
			return "safe"
		for Match in Result.PartialMatches:
			if not self._RecordMatchesElementType(Match.Record, ElementType):
				continue
			NonOptionalNodes = sum(1 for Record in Match.Record.IterPath() if not dgm_database.IsOptionalNode(Record.Node))
			if NonOptionalNodes >= 2:
				return "partial"
		return "unidentified"

	def FindElementTypeByCanonical(self, Canonical: str) -> Optional[PreprocessElementType]:
		Normalized = dgm_database.NormalizeText(Canonical)
		for ElementType in self.Rules.ElementTypes:
			if dgm_database.NormalizeText(ElementType.Canonical) == Normalized:
				return ElementType
		return None

	def _RecordMatchesElementType(self, Record: dgm_database.ElementRecord, ElementType: PreprocessElementType) -> bool:
		Canonical = dgm_database.NormalizeText(ElementType.Canonical)
		DisplayName = dgm_database.NormalizeText(Record.DisplayName)
		return DisplayName == Canonical or DisplayName.startswith(Canonical + " ")

	def GetStageName(self, StageId: str) -> str:
		for Stage in PREPROCESS_STAGES:
			if Stage.Id == StageId:
				return Stage.Name
		return "All stages"


	def _CleanWhitespace(self, Text: str) -> str:
		Result = Text
		if self.Rules.TrimWhitespace:
			Result = Result.strip()
		if self.Rules.CollapseWhitespace:
			Result = " ".join(Result.split())
		return Result

	def _NormalizeUkrainianLookalikes(self, Text: str, Notes: List[str]) -> str:
		Updated = Text.translate(UKRAINIAN_LOOKALIKE_TRANSLATION)
		if Updated != Text:
			Notes.append("Stage 1: normalized Latin/Russian lookalike letters to Ukrainian")
		return Updated

	def _ApplyRules(self, Text: str, Rules: Sequence[PreprocessRegexRule], StageName: str, Notes: List[str], ElementType: Optional[PreprocessElementType] = None) -> str:
		Current = Text
		for Rule in Rules:
			if not Rule.Enabled:
				continue
			if ElementType is not None and Rule.ElementTypes and ElementType.Id not in Rule.ElementTypes:
				continue
			try:
				Updated = re.sub(Rule.Pattern, ConvertReplacement(Rule.Replacement), Current, flags=self.RegexFlags)
			except re.error as Error:
				Notes.append(f"{StageName}: skipped invalid rule {Rule.Id}: {Error}")
				continue
			if Updated != Current:
				Notes.append(f"{StageName}: applied {Rule.Id}")
				Current = Updated
		return Current

	def _NormalizeType(self, Text: str, Notes: List[str]) -> Tuple[str, bool, bool]:
		Explicit = self._FindExplicitType(Text)
		if Explicit is not None:
			ElementType, Remainder = Explicit
			Updated = self._CleanWhitespace(f"{ElementType.Canonical} {Remainder}")
			if Updated != Text:
				Notes.append(f"Stage 2: moved/capitalized type as {ElementType.Canonical}")
			return Updated, self._DatabaseAccepts(ElementType, Updated), False

		for PrefixRule in self.RuntimePrefixRules:
			if PrefixRule.Pattern.search(Text) is None:
				continue
			ElementType = PrefixRule.ElementType
			Candidate = self._MakeTypedCandidate(ElementType, Text)
			if self._DatabaseAccepts(ElementType, Candidate):
				Notes.append(f"Stage 2: inferred {ElementType.Canonical} from database prefix pattern")
				return Candidate, True, False

		Matches: List[Tuple[PreprocessElementType, str]] = []
		for ElementType in self.Rules.ElementTypes:
			Candidate = self._MakeTypedCandidate(ElementType, Text)
			if self._DatabaseAccepts(ElementType, Candidate):
				Matches.append((ElementType, Candidate))

		if len(Matches) == 1:
			ElementType, Candidate = Matches[0]
			Notes.append(f"Stage 2: inferred {ElementType.Canonical} by database search")
			return Candidate, True, False
		if len(Matches) > 1:
			Notes.append("Stage 2: ambiguous database type inference")
			return Text, False, True
		return Text, False, False

	def _FindExplicitType(self, Text: str) -> Optional[Tuple[PreprocessElementType, str]]:
		for ElementType in self.Rules.ElementTypes:
			Aliases = [ElementType.Canonical] + ElementType.Aliases
			for Alias in sorted(Aliases, key=len, reverse=True):
				Pattern = re.compile(rf"(^|\s)({re.escape(Alias)})(\s|$)", flags=self.RegexFlags)
				Match = Pattern.search(Text)
				if Match is None:
					continue
				Remainder = self._CleanWhitespace((Text[:Match.start(2)] + " " + Text[Match.end(2):]).strip())
				return ElementType, Remainder
		return None

	def _FindLeadingExplicitType(self, Text: str) -> Optional[Tuple[PreprocessElementType, str]]:
		for ElementType in self.Rules.ElementTypes:
			Aliases = [ElementType.Canonical] + ElementType.Aliases
			for Alias in sorted(Aliases, key=len, reverse=True):
				Pattern = re.compile(rf"^\s*({re.escape(Alias)})(\s|$)", flags=self.RegexFlags)
				Match = Pattern.search(Text)
				if Match is None:
					continue
				Remainder = self._CleanWhitespace(Text[Match.end(1):])
				return ElementType, Remainder
		return None

	def _MakeTypedCandidate(self, ElementType: PreprocessElementType, Text: str) -> str:
		Candidate = self._CleanWhitespace(f"{ElementType.Canonical} {Text}")
		return self._ApplyTechnicalNormalization(Candidate, [])

	def _BuildRuntimePrefixRules(self) -> List[RuntimePrefixRule]:
		Rules: List[RuntimePrefixRule] = []
		ElementTypesByCanonical = {dgm_database.NormalizeText(ElementType.Canonical): ElementType for ElementType in self.Rules.ElementTypes}

		def Walk(Node: XmlTree.Element, OptionalParents: bool, ActiveType: Optional[PreprocessElementType]) -> None:
			NodeText = Node.get("text", "")
			NodeType = ElementTypesByCanonical.get(dgm_database.NormalizeText(NodeText), ActiveType)
			PatternText = Node.get(DATABASE_PREFIX_REGEX_ATTR, "").strip()
			NodePrefixText = NodeText.strip().rstrip("-–—")
			if OptionalParents and NodeType is not None and PatternText and len(NodePrefixText) >= 2:
				try:
					Rules.append(RuntimePrefixRule(NodeType, re.compile(PatternText, flags=self.RegexFlags)))
				except re.error:
					pass
			ChildrenHaveOptionalParents = OptionalParents and dgm_database.IsOptionalNode(Node)
			for Child in list(Node):
				if Child.tag == "node":
					Walk(Child, ChildrenHaveOptionalParents, NodeType)

		for Child in list(self.Database.CatalogNode):
			if Child.tag == "node":
				Walk(Child, True, None)
		return Rules

	def _DatabaseAccepts(self, ElementType: PreprocessElementType, Candidate: str) -> bool:
		if not ElementType.DatabaseCheck:
			return True
		Result = self.Database.FindElement(Candidate)
		Record = Result.Record
		return (
			Record is not None
			and Record.HasDgm
			and not Result.MatchedByRegex
			and self._RecordMatchesElementType(Record, ElementType)
		)

def BuildExactDatabaseText(Record: dgm_database.ElementRecord) -> str:
	Parts: List[str] = []
	for PathRecord in Record.IterPath():
		if not PathRecord.ConsumedText:
			continue
		if PathRecord.MatchedByRegex:
			Parts.append(PathRecord.ConsumedText)
		else:
			Parts.append(PathRecord.DisplayText)
	return "".join(Parts)


def LoadPreprocessRules(PathToRules: Path) -> PreprocessRules:
	if not PathToRules.exists():
		CreateDefaultPreprocessRules(PathToRules)

	try:
		Tree = XmlTree.parse(PathToRules)
	except XmlTree.ParseError as Error:
		raise RuntimeError(f"Cannot parse preprocessing XML '{PathToRules}': {Error}") from Error

	Root = Tree.getroot()
	if Root.tag != "xlsx_preprocess_rules":
		raise RuntimeError(f"Unexpected preprocessing XML root in '{PathToRules}': {Root.tag}")

	Rules = PreprocessRules(Path=PathToRules, Tree=Tree, Root=Root)
	Settings = Root.find("settings")
	if Settings is not None:
		RegexFlags = Settings.find("regex_flags")
		Whitespace = Settings.find("whitespace")
		if RegexFlags is not None:
			Rules.IgnoreCase = ReadBool(RegexFlags.get("ignore_case"), True)
		if Whitespace is not None:
			Rules.CollapseWhitespace = ReadBool(Whitespace.get("collapse"), True)
			Rules.TrimWhitespace = ReadBool(Whitespace.get("trim"), True)

	IgnoredTextsNode = Root.find("ignored_texts")
	if IgnoredTextsNode is not None:
		Rules.IgnoredTextRules.extend(ParseIgnoreRule(Node) for Node in IgnoredTextsNode.findall("pattern"))

	for Stage in Root.findall("stage"):
		StageName = Stage.get("name", "")
		ParsedRules = [ParseRegexRule(Node) for Node in Stage.findall("rule")]
		if StageName == "language_normalization" or Stage.get("id") == "1":
			Rules.StageOneRules.extend(ParsedRules)
		elif StageName == "technical_normalization" or Stage.get("id") == "3":
			Rules.StageThreeRules.extend(ParsedRules)

	ElementTypesNode = Root.find("element_types")
	if ElementTypesNode is not None:
		for TypeNode in ElementTypesNode.findall("type"):
			Canonical = TypeNode.get("canonical", "").strip()
			TypeId = TypeNode.get("id", Canonical).strip()
			if not Canonical:
				continue
			AliasesNode = TypeNode.find("aliases")
			DatabaseCheckNode = TypeNode.find("database_check")
			Aliases = [Child.text.strip() for Child in (AliasesNode.findall("alias") if AliasesNode is not None else []) if Child.text and Child.text.strip()]
			Rules.ElementTypes.append(
				PreprocessElementType(
					Id=TypeId,
					Canonical=Canonical,
					Aliases=Aliases,
					DatabaseCheck=ReadBool(DatabaseCheckNode.get("enabled") if DatabaseCheckNode is not None else None, True),
					XmlNode=TypeNode,
				)
			)

	return Rules


def ParseIgnoreRule(Node: XmlTree.Element) -> PreprocessRegexRule:
	DescriptionNode = Node.find("description")
	Description = Node.get("description", "")
	if DescriptionNode is not None and DescriptionNode.text is not None:
		Description = DescriptionNode.text
	return PreprocessRegexRule(
		Id=Node.get("id", "unnamed_ignore_pattern"),
		Pattern=(Node.get("regex") or Node.text or "").strip(),
		Replacement="",
		Description=Description,
		Enabled=ReadBool(Node.get("enabled"), True),
	)


def ParseRegexRule(Node: XmlTree.Element) -> PreprocessRegexRule:
	PatternNode = Node.find("pattern")
	ReplacementNode = Node.find("replacement")
	DescriptionNode = Node.find("description")
	ElementTypes = [Item.strip() for Item in Node.get("element_types", "").split(",") if Item.strip()]
	return PreprocessRegexRule(
		Id=Node.get("id", "unnamed_rule"),
		Pattern=(PatternNode.text if PatternNode is not None and PatternNode.text is not None else ""),
		Replacement=(ReplacementNode.text if ReplacementNode is not None and ReplacementNode.text is not None else ""),
		Description=(DescriptionNode.text if DescriptionNode is not None and DescriptionNode.text is not None else ""),
		Enabled=ReadBool(Node.get("enabled"), True),
		ElementTypes=ElementTypes,
	)


def SavePreprocessRules(Rules: PreprocessRules) -> None:
	IndentXml(Rules.Root)
	Rules.Tree.write(Rules.Path, encoding="utf-8", xml_declaration=True)


def CreateDefaultPreprocessRules(PathToRules: Path) -> None:
	PathToRules.parent.mkdir(parents=True, exist_ok=True)
	Root = XmlTree.Element("xlsx_preprocess_rules", {"version": RULES_VERSION})
	Settings = XmlTree.SubElement(Root, "settings")
	XmlTree.SubElement(Settings, "regex_flags", {"ignore_case": "true", "unicode": "true"})
	XmlTree.SubElement(Settings, "whitespace", {"collapse": "true", "trim": "true"})

	IgnoredTexts = XmlTree.SubElement(Root, "ignored_texts")
	AddIgnorePattern(IgnoredTexts, "board_plate_word", "Rows containing board/PCB word плат-а/и/і/у/... variations", r"\bплат(а|и|і|у|ою|ею|ах|ами|ам)?\b")
	AddIgnorePattern(IgnoredTexts, "generic_context_word_pair", "Rows containing at least two separate context words that cannot be an element designation", r"(?=.*\b(?:складі|стійці|розташовані|н[іи]й|ньому|якому|якій|яких|який|яка|яке|цей|цьому|цій)\b)\b(?:на|у|в|складі|стійці|розташовані|н[іи]й|ньому|якому|якій|яких|який|яка|яке|де|що|цей|ця|це|цьому|цій)\b.*\b(?:на|у|в|складі|стійці|розташовані|н[іи]й|ньому|якому|якій|яких|який|яка|яке|де|що|цей|ця|це|цьому|цій)\b")
	AddIgnorePattern(IgnoredTexts, "generic_missing_marker", "Rows containing explicit missing/excluded wording", r"\b(відсутн[іи]|немає|відсутнє|без)\b")
	AddIgnorePattern(IgnoredTexts, "generic_safe_ignore_words", "Rows containing safe non-element words крім/того/розміщені variants", r"\b(о?крім\w*|того|розміщен\w*)\b")
	AddIgnorePattern(IgnoredTexts, "generic_military_rank", "Rows containing Ukrainian military rank words", r"\b(рекрут\w*|солдат\w*|сержант\w*|старшин\w*|прапорщик\w*|лейтенант\w*|капітан\w*|майор\w*|підполковник\w*|полковник\w*|генерал\w*)\b")
	AddIgnorePattern(IgnoredTexts, "generic_commission", "Rows containing commission wording", r"\bкомісі\w*\b")
	AddIgnorePattern(IgnoredTexts, "generic_completeness", "Rows containing completeness/complectation wording", r"\b(комплектн\w*|комплектаці\w*)\b")

	AutogeneratedIgnoreAttributes = {"source": "autogenerated", "safety": "less_safe"}
	AutogeneratedIgnorePatterns = [
		("autogenerated_calculated_by", "Autogenerated less-safe ignored text: calculation signature row", r"(?iu)^\s*Підрахунок здійснили:.*$"),
		("autogenerated_available_property_heading", "Autogenerated less-safe ignored text: available property heading", r"(?iu)^\s*Найменування наявного майна та комплектуючих\s*$"),
		("autogenerated_documentation_reference", "Autogenerated less-safe ignored text: documentation reference", r"(?iu)^\s*згідно документації\s*$"),
		("autogenerated_blank_signature_date", "Autogenerated less-safe ignored text: blank signature date", r"(?iu)^\s*_{2,}\._{2,}\.\d{4}\s*$"),
		("autogenerated_empty_context_marker", "Autogenerated less-safe ignored text: empty context marker", r"(?iu)^\s*(?:Навісом|У\s*нь?ому|Уньому)\s*:?\s*$"),
		("autogenerated_block_rows", "Autogenerated less-safe ignored text: block rows", r"(?iu)^\s*(?:Блок|блок|БЖ)\b.*$"),
		("autogenerated_rack_rows", "Autogenerated less-safe ignored text: rack rows", r"(?iu)^\s*Стійк[аи]\b.*$"),
		("autogenerated_product_lab_rows", "Autogenerated less-safe ignored text: product/lab rows", r"(?iu)^\s*(?:Виріб|Лабораторія)\b.*$"),
		("autogenerated_cell_rows", "Autogenerated less-safe ignored text: cell/chamber rows", r"(?iu)^\s*(?:Комірка|Чарунка)\b.*$"),
		("autogenerated_computer_part_rows", "Autogenerated less-safe ignored text: computer part rows", r"(?iu)^\s*(?:Системний\s+блок|Монітор|Клавіатура|ОЗУ)\b.*$"),
		("autogenerated_storage_media_rows", "Autogenerated less-safe ignored text: storage/media rows", r"(?iu)^\s*(?:flo+p{1,2}y\s+drive|Дисковод|Привід\s+для\s+дискет|Електроника\s+МС5305\b.*\bFlop\b|Жорсткий\s+диск|Магнітофон)\b.*$"),
		("autogenerated_mechanical_optical_rows", "Autogenerated less-safe ignored text: mechanical/optical rows", r"(?iu)^\s*(?:алюм.*короб|Каркас\s+шафи|Корпус\s+з\b|механічна\s+система|Механізм\s+наведення|оптична\s+система|Насадка\s+бінокулярна|антена\b|Антена\b|телефонна\s+гарнітура)\b.*$"),
		("autogenerated_motor_display_rows", "Autogenerated less-safe ignored text: motor/display rows", r"(?iu)^\s*(?:Вентилятор|Двигун|АДП\s+\S+.*[эе]лектродв\.?|К[іи]н[еі]скоп|ЭЛТ)\b.*$"),
		("autogenerated_device_instrument_rows", "Autogenerated less-safe ignored text: device/instrument rows", r"(?iu)^\s*(?:Ампермети?р|Вольтмети?р|Мегоометр|Прибор\s+комбінований|Перетворювач|Пристрій\b|Пост\s+фотокінодеодолітний|Годинник|Часи|часи|Радіостанція|радіостанція)\b.*$"),
		("autogenerated_generic_element_counts", "Autogenerated less-safe ignored text: generic element-count rows", r"(?iu)^\s*(?:(?:[0-9]{2,4}(?:\s+[0-9]{2,4})?|[0-9]+[RК][0-9]+|I?R\d+|[N]|БН)\s+елемент|елемент(?:и)?\s+БН)\s*$"),
		("autogenerated_transformers_heading", "Autogenerated less-safe ignored text: transformers heading", r"(?iu)^\s*Трансформатори\s*$"),
	]
	for PatternId, Description, Pattern in AutogeneratedIgnorePatterns:
		AddIgnorePattern(IgnoredTexts, PatternId, Description, Pattern, AutogeneratedIgnoreAttributes)

	StageOne = XmlTree.SubElement(Root, "stage", {"id": "1", "name": "language_normalization"})
	AddRule(StageOne, "ru_relay_res", "Russian РЭС relay spelling to Ukrainian РЕС", r"\bРЭС\b", "РЕС")
	AddRule(StageOne, "ru_diode_to_uk_diode", "Russian диод to Ukrainian діод", r"\bдиод\b", "діод")

	Types = XmlTree.SubElement(Root, "element_types")
	AddElementType(Types, "capacitor", "Конденсатор", ["конденсатор", "конд", "кондер"])
	AddElementType(Types, "resistor", "Резистор", ["резистор", "сопротивление", "опір"])
	AddElementType(Types, "diode", "Діод", ["діод", "диод"])
	AddElementType(Types, "relay", "Реле", ["реле"])

	StageThree = XmlTree.SubElement(Root, "stage", {"id": "3", "name": "technical_normalization"})
	AddRule(StageThree, "km_series_hyphenation", "Normalize КМ capacitor names like КМ 5б Н90 to КМ-5б-Н90", r"\bКМ[\s\-]*(\d+[а-яА-Яa-zA-Z]?)[\s\-]*(Н\d+)\b", "КМ-$1-$2", "capacitor")
	AddRule(StageThree, "capacitor_k_prefix_space_hyphenation", "Normalize К-series capacitor designations with a space after К, e.g. К 50 6 to К50-6", r"\b[КK]\s+(10|15|22|31|40|41|42|50|52|53|57|71|72|73|75|76|77|78)[\s\-]+(\d+[а-яА-Яa-zA-Z]*)\b", "К$1-$2", "capacitor")
	AddRule(StageThree, "resistor_sp_prefix_space_hyphenation", "Normalize СП resistor designations with split СП/series groups, e.g. СП 5 16 to СП5-16", r"\b[СC][\s\-]*[ПP]\s*(\d+)[\s\-]+(\d+[а-яА-Яa-zA-Z]*)\b", "СП$1-$2", "resistor")
	AddRule(StageThree, "resistor_s_prefix_space_hyphenation", "Normalize С-series resistor designations with a space after С, e.g. С 2 14 to С2-14", r"\b[СC]\s+(1|2|3|5|6)[\s\-]+(\d+[а-яА-Яa-zA-Z]*)\b", "С$1-$2", "resistor")
	AddRule(StageThree, "diode_2d_prefix_join", "Join 2Д diode prefixes split by whitespace", r"\b2\s*[ДD][\s\-]*(\d+[а-яА-Яa-zA-Z]*)\b", "2Д$1", "diode")
	AddRule(StageThree, "diode_2s_prefix_join", "Join 2С diode/stabilizer prefixes split by whitespace", r"\b2\s*[СC][\s\-]*(\d+[а-яА-Яa-zA-Z]*)\b", "2С$1", "diode")
	AddRule(StageThree, "diode_2a_prefix_join", "Join 2А diode prefixes split by whitespace", r"\b2\s*[АA][\s\-]*(\d+[а-яА-Яa-zA-Z]*)\b", "2А$1", "diode")
	AddRule(StageThree, "diode_3i_prefix_join", "Join 3И diode prefixes split by whitespace", r"\b3\s*[ИI][\s\-]*(\d+[а-яА-Яa-zA-Z]*)\b", "3И$1", "diode")
	AddRule(StageThree, "diode_d_series_join", "Normalize common Д-series diode designations split by whitespace", r"\b[ДD][\s\-]+(\d{1,4})\b", "Д$1", "diode")
	AddRule(StageThree, "diode_d_trailing_letter_join", "Remove accidental space before the trailing letter in Д-series diode designations", r"\b(Д\d{1,4})\s+([а-яА-Яa-zA-Z])\b", "$1$2", "diode")
	AddRule(StageThree, "semiconductor_trailing_letter_join", "Remove accidental space before trailing letters in common semiconductor designations", r"\b((?:КД|КС|КЦ|АЛ|КТ|ГТ|МП|КП|КН|КУ|П)\d{1,4})\s+([а-яА-Яa-zA-Z])\b", "$1$2", "diode,transistor,dinistor,thyristor")
	AddRule(StageThree, "transistor_numeric_t_prefix_join", "Join numeric Т transistor prefixes, e.g. 2 Т 312 to 2Т312", r"\b([12])\s*[ТT][\s\-]*(\d+[а-яА-Яa-zA-Z]*)\b", "$1Т$2", "transistor")
	AddRule(StageThree, "thyristor_2u_prefix_join", "Join 2У thyristor prefixes split by whitespace", r"\b2\s*[УY][\s\-]*(\d+[а-яА-Яa-zA-Z]*)\b", "2У$1", "thyristor")
	AddRule(StageThree, "microchip_prefix_space_join", "Remove accidental space after microchip prefix, e.g. к 155ИД3 to к155ИД3", r"\b([КK][РPМM]?|[КK][МM])\s+(\d{3,4}[А-ЯA-Z]{1,4}\d{1,3}[А-ЯA-Z]?)\b", "$1$2", "microchip")
	AddRule(StageThree, "res_relay_series_hyphenation", "Normalize РЕС relay names with missing hyphen", r"\bРЕС[\s\-]*(\d+)\b", "РЕС-$1", "relay")

	Tree = XmlTree.ElementTree(Root)
	IndentXml(Root)
	Tree.write(PathToRules, encoding="utf-8", xml_declaration=True)


def AddIgnorePattern(Parent: XmlTree.Element, PatternId: str, Description: str, Pattern: str, ExtraAttributes: Optional[Dict[str, str]] = None) -> None:
	Attributes = {"id": PatternId, "enabled": "true", "description": Description}
	if ExtraAttributes:
		Attributes.update(ExtraAttributes)
	PatternNode = XmlTree.SubElement(Parent, "pattern", Attributes)
	PatternNode.text = Pattern


def AddRule(Parent: XmlTree.Element, RuleId: str, Description: str, Pattern: str, Replacement: str, ElementTypes: str = "") -> None:
	Attributes = {"id": RuleId, "enabled": "true"}
	if ElementTypes:
		Attributes["element_types"] = ElementTypes
	Rule = XmlTree.SubElement(Parent, "rule", Attributes)
	XmlTree.SubElement(Rule, "description").text = Description
	XmlTree.SubElement(Rule, "pattern").text = Pattern
	XmlTree.SubElement(Rule, "replacement").text = Replacement


def AddElementType(Parent: XmlTree.Element, TypeId: str, Canonical: str, Aliases: Sequence[str]) -> None:
	TypeNode = XmlTree.SubElement(Parent, "type", {"id": TypeId, "canonical": Canonical})
	AliasesNode = XmlTree.SubElement(TypeNode, "aliases")
	for Alias in Aliases:
		XmlTree.SubElement(AliasesNode, "alias").text = Alias
	XmlTree.SubElement(TypeNode, "database_check", {"enabled": "true"})


def ReadBool(Value: Optional[str], Default: bool) -> bool:
	if Value is None:
		return Default
	return Value.strip().casefold() in ("1", "true", "yes", "on")


def ConvertReplacement(Replacement: str) -> str:
	return re.sub(r"\$(\d+)", r"\\\1", Replacement)


def IndentXml(Element: XmlTree.Element, Level: int = 0) -> None:
	Indent = "\n" + Level * "\t"
	if len(Element):
		if not Element.text or not Element.text.strip():
			Element.text = Indent + "\t"
		for Child in Element:
			IndentXml(Child, Level + 1)
		if not Child.tail or not Child.tail.strip():
			Child.tail = Indent
	if Level and (not Element.tail or not Element.tail.strip()):
		Element.tail = Indent


def NormalizeDesignationCase(Text: str) -> str:
	def UpperLettersBeforeDigits(Match: re.Match[str]) -> str:
		return Match.group(1).upper() + Match.group(2)

	return re.sub(r"(?<!\d)([A-Za-zА-Яа-яІіЇїЄєҐґ]+)(\d+)", UpperLettersBeforeDigits, Text)
